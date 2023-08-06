#!/usr/bin/env python
from collections import Counter
import datetime
from datetime import datetime, timedelta
from multiprocessing import freeze_support
from re import MULTILINE
from colorama import init
from easydict import EasyDict as edict
from prophet.plot import plot_plotly
from jenkinsapi.jenkins import Jenkins
from json2html import *
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from argparse import ArgumentParser
from base64 import b64encode
from tempfile import gettempdir
from time import time
from unicodedata import normalize
from urllib import request
from numpy import NaN
from pandas import to_datetime, DataFrame, crosstab, read_csv, read_excel
from plotly import io
from plotly.express import pie, histogram
from requests import get
from tzlocal import get_localzone
from shutil import copy2
from html import escape
import json
import os
import platform
import re
import ssl
import subprocess
import sys
import glob

""" Microsoft Visual C++ required, cython required for pandas installation, """
TEMP_DIR = "/tmp" if platform.system() == "Darwin" else gettempdir()
live_report_filename = "Interactive_Report.html"
email_report_filename = "email.html"
orchestrationIssues = ["already in use"]
labIssues = ["HANDSET_ERROR", "No device was found"]
# Do not change these variable
RESOURCE_TYPE = "handsets"
RESOURCE_TYPE_USERS = "users"
REPOSITORY_RESOURCE_TYPE = "repositories/media"

addInfo, tags, reportTag, criteria, jobName, jobNumber, startDate, endDate, consolidate, port, temp, issues_email, title = "", "", "", "", "", "", "", "", "", "", "", "", ""
trends, report, debug = "false", "Report: ", "false"
cleanedFailureList, suggesstionsDict, issues, recommendations, execution_status = {}, {}, {}, {}, {}

"""
   sends request
"""


def send_request(url):
    """send request"""
    device_list_parameters = os.environ["DEVICE_LIST_PARAMETERS"]
    if (
        "All devices" in device_list_parameters
        or "Available devices only" in device_list_parameters
    ):
        response = request.urlopen(url)
    else:
        response = request.urlopen(url.replace(" ", "%20"))
    #    rc = response.getcode()
    return response


"""
   returns as text if none
"""


def as_text(value):
    """as texts"""
    if value is None:
        return ""
    return str(value)


"""
   validate logo
"""


def validate_logo(logo):
    try:
        send_request(logo)
    except Exception as e:
        print("Exception: " + str(e))
        os.environ["company_logo"] = os.environ["perfecto_logo"]


"""
  create pie
"""


def create_pie(df, title, column, name):
    status = []
    for i in range(len(df[column].value_counts().sort_index().to_frame())):
        status.append(df[column].value_counts(
        ).sort_index().to_frame().iloc[i].name)
    status_df = df['status'].value_counts().sort_index().to_frame()
    fig = pie(status_df, values=status_df['status'], names=status, color=status, opacity=1, hole=0.5, width=490, height=200,
              color_discrete_map={'PASSED': '#35a600',
                                  'FAILED': '#f14c4c',
                                  'BLOCKED': '#4cb2ff',
                                  'UNKNOWN': '#929da5'})
    fig.update_layout(margin=dict(t=0, b=0, l=0, r=0))
    img = fig.to_image(format="png")
    summary = '<img src="data:image/png;base64, {}"'.format(
        b64encode(img).decode("utf-8"))
    return summary


"""
    Dictionary
"""


class my_dictionary(dict):
    def __init__(self):
        self = dict()

    def add(self, key, value):
        self[key] = value


"""
    Creates payload for reporting API
"""


def payloadJobAll(reportTags, oldmilliSecs, current_time_millis, jobName, jobNumber, page, boolean):
    payload = my_dictionary()
    if oldmilliSecs != 0:
        payload.add("startExecutionTime[0]", oldmilliSecs)
    if reportTags != "":
        for i, reportTaging in enumerate(reportTags.split(";")):
            payload.add("tags[" + str(i) + "]", reportTaging)
    if current_time_millis != 0:
        payload.add("endExecutionTime[0]", current_time_millis)
    payload.add("_page", page)
    if jobName != "":
        if jobName != "All Jobs":
            if jobName != "Perfecto Integrations":
                for i, job in enumerate(jobName.split(";")):
                    payload.add("jobName[" + str(i) + "]", job)
    if jobNumber != "" and boolean:
        for i, jobNumber in enumerate(jobNumber.split(";")):
            payload.add("jobNumber[" + str(i) + "]", int(jobNumber))
    if debug == "true":
        print(str(payload))  # debug
    return payload


"""
    Retrieve a list of test executions within the last month
    :return: JSON object contains the executions
"""


def retrieve_tests_executions(daysOlder, page):
    current_time_millis = 0
    oldmilliSecs = 0
    global endDate
    if endDate != "":
        if "-" not in endDate:
            current_time_millis = endDate
        else:
            endTime = datetime.strptime(
                str(endDate) + " 23:59:59,999", "%Y-%m-%d %H:%M:%S,%f"
            )
            if debug == "true":
                print("endExecutionTime: " + str(endTime))  # debug
            millisec = endTime.timestamp() * 1000
            current_time_millis = round(int(millisec))
    if startDate != "":
        if "-" not in startDate:
            oldmilliSecs = startDate
        else:
            oldmilliSecs = pastDateToMS(startDate, daysOlder)
    global reportTag
    if jobNumber != "" and jobName != "" and startDate != "" and endDate != "":
        payload = payloadJobAll(
            reportTag, oldmilliSecs, current_time_millis, jobName, jobNumber, page, False
        )
    else:
        payload = payloadJobAll(
            reportTag, oldmilliSecs, current_time_millis, jobName, jobNumber, page, True
        )
    url = "https://" + os.environ["cloudName"] + \
        ".reporting.perfectomobile.com"
    api_url = url + "/export/api/v1/test-executions"
    # creates http geat request with the url, given parameters (payload) and header (for authentication)
    r = get(
        api_url, params=payload, headers={
            "PERFECTO_AUTHORIZATION": os.environ["securityToken"]}
    )
    if debug == "true":
        print(str(r.url))  # debug
    return r.content


def df_formatter(df):
    if len(df) < 1:
        raise Exception("Unable to find any matching executions!")
    try:
        df["startTime"] = to_datetime(df["startTime"], unit="ms")
        df["startTime"] = (
            df["startTime"].dt.tz_localize(
                "utc").dt.tz_convert(get_localzone())
        )
        df["startTime"] = df["startTime"].dt.strftime("%d/%m/%Y %H:%M:%S")
    except:
        pass
    try:
        df.loc[df["endTime"] < 1, "endTime"] = int(round(time() * 1000))
        df["endTime"] = to_datetime(df["endTime"], unit="ms")
        df["endTime"] = (
            df["endTime"].dt.tz_localize(
                "utc").dt.tz_convert(get_localzone())
        )
        df["endTime"] = df["endTime"].dt.strftime("%d/%m/%Y %H:%M:%S")
    except:
        pass
    if "month" not in df.columns:
        df["month"] = to_datetime(
            df["startTime"], format="%d/%m/%Y %H:%M:%S"
        ).dt.to_period("M")
    if "startDate" not in df.columns:
        df["startDate"] = to_datetime(
            to_datetime(df["startTime"], format="%d/%m/%Y %H:%M:%S")
            .dt.to_period("D")
            .astype(str)
        )
    if "week" not in df.columns:
        df["week"] = to_datetime(df["startDate"].dt.strftime("%Y/%m/%d")) - df[
            "startDate"
        ].dt.weekday.astype("timedelta64[D]")
    if "Duration" not in df.columns:
        df["Duration"] = to_datetime(df["endTime"]) - to_datetime(
            df["startTime"]
        )
        df["Duration"] = df["Duration"].dt.seconds
        df["Duration"] = to_datetime(df["Duration"], unit="s").dt.strftime(
            "%H:%M:%S"
        )
    if "failureReasonName" not in df.columns:
        df["failureReasonName"] = ""
    # df["name"] = '=HYPERLINK("'+df["reportURL"]+'", "'+df["name"]+'")'  # has the ability to hyperlink name in csv'
    # Filter only job and job number if dates are parameterized as well but show full histogram
    if jobNumber != "" and jobName != "":
        if ";" in jobNumber:
            df = df[df["job/number"].isin(jobNumber.split(";"))]
        else:
            df = df[df["job/number"].astype(str) == jobNumber]
    if startDate != "":
        name = startDate
    else:
        name = jobName + "_" + jobNumber
    if debug == "true":
        print("Actual count of dataframe: " + str(len(df)))
    df = df_to_xl(df, str(name).replace("/", "_"))
    return df


"""
    flattens the json
"""


def flatten_json(nested_json, exclude=[""]):
    """Flatten json object with nested keys into a single level.
        Args:
            nested_json: A nested json object.
            exclude: Keys to exclude from output.
        Returns:
            The flattened json object if successful, None otherwise.
    """
    out = {}

    def flatten(x, name="", exclude=exclude):
        if type(x) is dict:
            for a in x:
                if a not in exclude:
                    flatten(x[a], name + a + "/")
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + "/")
                i += 1
        else:
            out[name[:-1]] = x
    flatten(nested_json)
    return out


"""
    get final dataframe
"""


def get_final_df(files):
    df = DataFrame()
    for file in files:
        if "csv" in os.environ["xlformat"]:
            df = df.append(read_csv(file, low_memory=False))
        else:
            df = df.append(read_excel(file))
        print("Analysing file: " + str(file) + " , row count: " + str(len(df)))
    df = df_formatter(df)
    return df


"""
   calculates the percetage of a part and whole number
"""


def percentageCalculator(part, whole):
    if int(whole) > 0:
        calc = (100 * float(part) / float(whole), 0)
        calc = round(float((calc[0])), 2)
    else:
        calc = 0
    return calc


"""
   gets start date to milliseconds
"""


def pastDateToMS(startDate, daysOlder):
    dt_obj = datetime.strptime(
        startDate + " 00:00:00,00", "%Y-%m-%d %H:%M:%S,%f"
    ) - timedelta(days=daysOlder)
    millisec = dt_obj.timestamp() * 1000
    oldmilliSecs = round(int(millisec))
    return oldmilliSecs


"""
   creates graphs
"""


def createGraphs(interactive_graphs, graphs, duration, df, height, width, jobName, job, counter):
        predict_df = df
        fig = []
        if job != "Overall!":
            if job in jobName.split(";") or jobName == "All Jobs" or str(job).endswith("Sample"):
                if duration == "dates":
                    fig = histogram(
                        df.loc[df["job/name"] == job],
                        x="startDate",
                        color="status",
                        color_discrete_map={
                            "PASSED": "limegreen",
                            "FAILED": "crimson",
                            "UNKNOWN": "#9da7f2",
                            "BLOCKED": "#e79a00",
                        },
                        hover_data=df.columns,
                        template="seaborn",
                        opacity=0.5,
                        height=height,
                        width=width
                    )
                else:
                    fig = histogram(
                        df.loc[df["job/name"] == job],
                        x="week",
                        color="status",
                        hover_data=df.columns,
                        color_discrete_map={
                            "PASSED": "limegreen",
                            "FAILED": "crimson",
                            "UNKNOWN": "#9da7f2",
                            "BLOCKED": "#e79a00",
                        },
                        template="seaborn",
                        opacity=0.5,
                        height=height,
                        width=width
                    )
                predict_df = df.loc[df["job/name"] == job]
        else:
            fig = histogram(
                df,
                x="startDate",
                color="status",
                color_discrete_map={
                    "PASSED": "limegreen",
                    "FAILED": "crimson",
                    "UNKNOWN": "#9da7f2",
                    "BLOCKED": "#e79a00",
                },
                hover_data=df.columns,
                template="seaborn",
                opacity=0.5,
                height=height,
                width=width
            )
        predict_df = (
            predict_df.groupby(["startDate"])
            .size()
            .reset_index(name="#status")
            .sort_values("#status", ascending=False)
        )
        tabcontent = 'style="box-sizing: border-box; display: contents; height: auto; -moz-transition: height 1s ease; -webkit-transition: height 1s ease; -o-transition: height 1s ease; transition: height 1s ease; overflow: auto; justify-content: center;"'
        reportDiv = (
            'style="box-sizing: border-box; overflow-x: auto; text-align: -webkit-center;"'
        )
        graphPane = 'style="overflow-x: auto;text-align: center;display: inline-block;background-color:white;width:100%;padding-left: 14%;"'
        predictionDiv = 'style="overflow-x: auto;text-align: center;display: inline-block;float: left;background-color:white;width:100%;"'
        header = 'align=center; style="box-sizing: border-box; float: left; width: 100%; padding: 1px 0; text-align: center; cursor: pointer; font-size: 16px; color: darkslategray; background-color: darkkhaki; border: 3px solid antiquewhite;"'
        if fig:
            ci_name = os.environ["ci_name"]
            tag = ""
            if ci_name != "":
                ci_jenkins_url = os.environ["ci_jenkins_url"]
                ci_username = os.environ["ci_username"]
                ci_token = os.environ["ci_token"]
                ci_href = ""
                ci_server_url = ""
                ci_src = ""
                if "jenkins" in ci_name.lower():
                    circleCIjobs = [
                        "MavenCircleCISample", "fastlane-plugin-perfecto", "FastlaneEspressoCircleCISlackSample"]
                    travisCIjobs = ["TravisSample"]
                    group = "PerfectoMobileSA"
                    if job in circleCIjobs:
                        ci_server_url = "https://circleci.com/gh/" + group
                        ci_href = ci_server_url + "/" + job
                        ci_src = ci_server_url + "/" + job + ".svg?style=shield"
                    elif job in travisCIjobs:
                        ci_server_url = "https://travis-ci.org/" + group
                        ci_href = ci_server_url + "/" + job
                        ci_src = ci_server_url + "/" + job + ".svg?branch=master"
                    else:
                        j = Jenkins(
                            ci_jenkins_url, username=ci_username, password=ci_token)
                        ci_server_url = str(j).split("at ")[1]
                        ci_href = str(ci_server_url + "/job/" + job)
                        ci_src = ci_server_url + "/buildStatus/icon?job=" + job
                    print("ci_server_url: " + ci_server_url +
                            ", ci_href: " + ci_href + ", ci_src=" + ci_src)
                    tag = " <a href='" + ci_href + "' target='_blank'><img src='" + ci_src + "'></a>"
            fig = update_fig(fig, "histogram", job, duration)
            encoded = b64encode(io.to_image(fig))
            graphs.append('<div '
                            + header
                            + '><b><center>job: '
                            + job
                            + tag
                            + '</center></b></label></div><div align="center" class="tab-content1" '
                            + tabcontent
                            + "><div " + predictionDiv + ">"
                            + '<img src="data:image/png;base64, {}"'.format(
                                encoded.decode("ascii")
                            )
                            + " alt='days or weeks summary of "
                            + job
                            + "' id='reportDiv' "
                            + reportDiv
                            + "> </img>"
                            )
            interactive_graphs.append(
                '<div style="text-align: center;"><input type="radio" id="tab'
                + str(counter)
                + '" name="tabs" checked=""/><label for="tab'
                + str(counter)
                + '">job: '
                + job
                + tag
                + '</label><div class="tab-content1"'
                + tabcontent
                + "><div " + graphPane + ">"
                + str(fig.to_html(full_html=False, include_plotlyjs="cdn")
                        ).replace("<div", '<div style="float:left;"')
            )
        if job == "Overall!" or job in jobName or jobName == "All Jobs" or str(jobName).endswith("Sample"):
            if job in jobName or jobName == "All Jobs":
                if len(predict_df.index) > 1:
                    predict_df = predict_df.rename(
                        columns={"startDate": "ds", "#status": "y"}
                    )
                    predict_df["cap"] = int(predict_df["y"].max()) * 2
                    predict_df["floor"] = 0
                    from prophet import Prophet

                    with suppress_stdout_stderr():
                        m = Prophet(
                            seasonality_mode="additive",
                            growth="logistic",
                            changepoint_prior_scale=0.3,
                            weekly_seasonality=False,
                            daily_seasonality=True,
                            yearly_seasonality=False,
                        ).fit(predict_df, algorithm="Newton")
                    future = m.make_future_dataframe(periods=30, freq='d')
                    future["cap"] = int(predict_df["y"].max()) * 1.6
                    floor = 0
                    if (int(predict_df["y"].min()) / 2) > 0:
                        floor = int(predict_df["y"].min())
                    future["floor"] = floor
                    forecast = m.predict(future)
                    forecast[["ds", "yhat", "yhat_lower", "yhat_upper"]].tail()
                    fig = plot_plotly(
                        m, forecast, figsize=([height, width]))
                    fig = update_fig(fig, "prediction", job, duration)
                    encoded = b64encode(io.to_image(fig))
                    graphs.append('<div '
                                    + predictionDiv
                                    + '>'
                                    + '<img src="data:image/png;base64, {}"'.format(
                                        encoded.decode("ascii")
                                    )
                                    + " alt='prediction of "
                                    + job
                                    + "' id='reportDiv' "
                                    + reportDiv
                                    + "> </img></div><br>"
                                    )
                    interactive_graphs.append(
                        '<div class="predictionDiv">'
                        + fig.to_html(full_html=False,
                                        include_plotlyjs="cdn")
                        + " </img></p></div>"
                    )
                    interactive_graphs.append("</div>")
                else:
                    interactive_graphs.append("</div>")
                    if debug == "true":
                        print(
                            "Note: AI Prediction for job: "
                            + job
                            + " requires executions of more than 2 days!"

                        )
        counter += 1   
        return interactive_graphs, graphs
                
"""
   gets' Perfecto reporting API responses, creates dict for top device failures, auto suggestions and top tests failures and prepared json
"""


def prepareReport(jobName, jobNumber, reportTag):
    page = 1
    i = 0
    truncated = True
    resources = []
    print("#Parameters:")
    print("endDate: " + endDate)
    print("startDate: " + startDate)
    print("jobName: " + jobName)
    print("jobNumber: " + jobNumber)
    print("tags: " + reportTag)
    if debug == "true":
        json_raw = os.environ["cloudName"] + "_API_output" + '.txt'
        open(json_raw, 'w').close
    print("Downloading test executions from Smart Reporting!")
    while truncated == True:
        if debug == "true":
            print(
                "Downloading test executions from Smart Reporting. Current page: "
                + str(page)
            )
        executions = retrieve_tests_executions(0, page)
        # print(executions)
        # Loads JSON string into JSON object
        executions = json.loads(executions)
        if "{'userMessage': 'Failed decoding the offline token:" in str(executions):
            raise Exception("please change the offline token for your cloud")
        if "userMessage': 'Missing Perfecto-TenantId header" in str(executions):
            raise Exception("Check the cloud name and security tokens")
        if "userMessage': 'Time period is not in supported range" in str(executions):
            raise Exception(
                "Time period is not in supported range. Check your startDate parameter")
        try:
            executionList = executions["resources"]
        except TypeError:
            print(executions)
            raise Exception(
                "Unable to find matching records for: "
                + str(criteria)
                + ", error:"
                + str(executions["userMessage"])
            )
            sys.exit(-1)
        if len(executionList) == 0:
            print("0 test executions")
            break
        else:
            # print(str(executions))
            metadata = executions["metadata"]
            truncated = metadata["truncated"]
            if page >= 1:
                resources.extend(executionList)
            else:
                resources.append(executionList)
            page += 1
    if len(resources) > 0:
        jsonDump = json.dumps(resources)
        resources = json.loads(jsonDump)
        if debug == "true":
            with open(json_raw, "a", encoding="utf-8") as myfile:
                myfile.write(str(resources) +
                             '\n*******************************************\n')
        print("Total executions: " + str(len(resources)))
        df = DataFrame([flatten_json(x) for x in resources])
        df = df_formatter(df)
    os.chdir(".")
    files = glob.glob("*.{}".format(os.environ["xlformat"]))
    consolidate = os.environ["consolidate"]
    if consolidate != "":
        for file in files:
            if os.path.isfile(file):
                copy2(file, consolidate)
        files = glob.iglob(os.path.join(
            consolidate, "*." + os.environ["xlformat"]))
    df = get_final_df(files)
    df = df.sort_values(by=["startDate"], ascending=False)
    if jobNumber != "" and jobName != "":
        if jobName != "All Jobs":
            if jobName != "Perfecto Integrations":
                df = df[df["job/name"].astype(str).isin(jobName.split(";"))]
                df = df[df["job/number"].round(0).astype(
                    str).isin(jobNumber.split(";"))]
    if jobNumber == "" and jobName != "":
        if jobName != "All Jobs":
            if jobName != "Perfecto Integrations":
                df = df[df["job/name"].astype(str).isin(jobName.split(";"))]
    # No support for tags in consolidation
    # if reportTag != "":
    #     l = [tuple(i) for i in reportTag.split(";")]
    #     df = df[df[df.columns[pandas.Series(df.columns).str.startswith('tags/')]].apply(tuple, axis = 1).astype(str).isin(l)]
    df = df_to_xl(df, "final")
    if (len(df)) < 1:
        print("Unable to find any test executions for expected parameters")
        sys.exit(-1)

    # ggplot2 #plotly_dark #simple_white
    graphs = []
    interactive_graphs = []
    width = 400
    height = 400
    if trends == "true":
        counter = 7
        with open(live_report_filename, "a") as f:
            f.write('<div id="nestle-section">')
        duration = "weeks"
        if startDate != "":
            delta = datetime.strptime(endDate, "%Y-%m-%d") - datetime.strptime(
                startDate, "%Y-%m-%d"
            )
            if (delta.days) <= 14:
                duration = "dates"
        else:
            duration = "dates"
        
        joblist = []
        if "job/name" in df.columns and jobName != "":
            joblist = sorted(df["job/name"].dropna().unique())
        else:
            joblist.append("Overall!")
        # Don't try to pool this process as its consuming more time to setup orca
        for job in joblist:
            interactive_graphs, graphs = createGraphs(interactive_graphs, graphs, duration, df, height, width, jobName, job, counter)
        graphs.append("</div>")
        interactive_graphs.append("</div>")
        with open(live_report_filename, "a") as f:
            f.write("</div>") 
    return graphs, interactive_graphs, df


"""
   suppress prophet logs
"""


class suppress_stdout_stderr(object):
    """
    A context manager for doing a "deep suppression" of stdout and stderr in
    Python, i.e. will suppress all print, even if the print originates in a
    compiled C/Fortran sub-function.
       This will not suppress raised exceptions, since exceptions are printed
    to stderr just before a script exits, and after the context manager has
    exited (at least, I think that is why it lets exceptions through).

    """

    def __init__(self):
        # Open a pair of null files
        self.null_fds = [os.open(os.devnull, os.O_RDWR) for x in range(2)]
        # Save the actual stdout (1) and stderr (2) file descriptors.
        self.save_fds = (os.dup(1), os.dup(2))

    def __enter__(self):
        # Assign the null pointers to stdout and stderr.
        os.dup2(self.null_fds[0], 1)
        os.dup2(self.null_fds[1], 2)

    def __exit__(self, *_):
        # Re-assign the real stdout/stderr back to (1) and (2)
        os.dup2(self.save_fds[0], 1)
        os.dup2(self.save_fds[1], 2)
        # Close the null files
        os.close(self.null_fds[0])
        os.close(self.null_fds[1])


"""
   converts datafame to excel
"""


def df_to_xl(df, filename):
    custom_columns = [
        "name",
        "status",
        "platforms/0/os",
        "platforms/0/mobileInfo/model",
        "platforms/0/browserInfo/browserType",
        "platforms/0/browserInfo/browserVersion",
        "platforms/0/osVersion",
        "failureReasonName",
        "message",
        "startTime",
        "endTime",
        "Duration",
        "job/name",
        "job/number",
        "job/branch",
        "owner",
        "reportURL",
        "platforms/0/deviceId",
        "platforms/0/deviceType",
        "platforms/0/mobileInfo/manufacturer",
        "platforms/0/screenResolution",
        "platforms/0/location",
        "platforms/0/mobileInfo/imei",
        "platforms/0/mobileInfo/phoneNumber",
        "platforms/0/mobileInfo/distributor",
        "platforms/0/mobileInfo/firmware",
        "platforms/0/selectionCriteriaV2/0/name",
        "platforms/0/selectionCriteriaV2/1/name",
        "platforms/0/selectionCriteriaV2/2/name",
        "platforms/0/selectionCriteriaV2/2/value",
        "platforms/0/customFields/0/name",
        "platforms/0/customFields/0/value",
        "tags/0",
        "tags/1",
        "tags/2",
        "tags/3",
        "tags/4",
        "tags/5",
        "tags/6",
        "tags/7",
        "tags/8",
        "tags/9",
        "tags/10",
        "tags/11",
        "tags/12",
        "tags/13",
        "tags/14",
        "tags/15",
        "tags/16",
        "tags/17",
        "tags/18",
        "tags/19",
        "tags/20",
        "id",
        "externalId",
        "uxDuration",
        "videos/0/startTime",
        "videos/0/endTime",
        "videos/0/format",
        "videos/0/streamingUrl",
        "videos/0/downloadUrl",
        "videos/0/screen/width",
        "videos/0/screen/height",
        "executionEngine/version",
        "project/name",
        "project/version",
        "automationFramework",
        "parameters/0/name",
        "parameters/0/value",
        "parameters/1/name",
        "parameters/1/value",
        "parameters/2/name",
        "parameters/2/value",
        "parameters/3/name",
        "parameters/3/value",
        "parameters/4/name",
        "parameters/4/value",
        "parameters/5/name",
        "parameters/5/value",
        "parameters/6/name",
        "parameters/6/value",
        "parameters/7/name",
        "parameters/7/value",
        "parameters/8/name",
        "parameters/8/value",
        "parameters/9/name",
        "parameters/9/value",
        "parameters/10/name",
        "parameters/10/value",
        "parameters/11/name",
        "parameters/11/value",
        "parameters/12/name",
        "parameters/12/value",
        "parameters/13/name",
        "parameters/13/value",
        "platforms/0/mobileInfo/operator",
        "platforms/0/mobileInfo/operatorCountry",
        "platforms/0/selectionCriteriaV2/3/name",
        "platforms/0/selectionCriteriaV2/3/value",
        "platforms/0/selectionCriteriaV2/4/name",
        "platforms/0/selectionCriteriaV2/4/value",
        "platforms/0/selectionCriteriaV2/5/name",
        "platforms/0/selectionCriteriaV2/5/value",
        "platforms/0/selectionCriteriaV2/6/name",
        "platforms/0/selectionCriteriaV2/6/value",
        "platforms/0/selectionCriteriaV2/7/name",
        "platforms/0/selectionCriteriaV2/7/value",
        "customFields/0/name",
        "customFields/0/value",
        "customFields/1/name",
        "customFields/1/value",
        "artifacts/0/type",
        "artifacts/0/path",
        "artifacts/0/zipped",
        "artifacts/1/type",
        "artifacts/1/path",
        "artifacts/1/contentType",
        "artifacts/1/zipped",
        "artifacts/2/type",
        "artifacts/2/path",
        "artifacts/2/zipped",
        "artifacts/0/contentType",
        "artifacts/2/contentType",
        "platforms/1/deviceId",
        "platforms/1/deviceType",
        "platforms/1/os",
        "platforms/1/osVersion",
        "platforms/1/screenResolution",
        "platforms/1/location",
        "platforms/1/mobileInfo/imei",
        "platforms/1/mobileInfo/manufacturer",
        "platforms/1/mobileInfo/model",
        "platforms/1/mobileInfo/distributor",
        "platforms/1/mobileInfo/firmware",
        "platforms/1/selectionCriteriaV2/0/name",
        "platforms/1/selectionCriteriaV2/0/value",
        "platforms/1/customFields/0/name",
        "platforms/1/customFields/0/value",
        "videos/1/startTime",
        "videos/1/endTime",
        "videos/1/format",
        "videos/1/streamingUrl",
        "videos/1/downloadUrl",
        "videos/1/screen/width",
        "videos/1/screen/height",
        "platforms/1/mobileInfo/phoneNumber",
        "month",
        "week",
        "startDate",
    ]
    df = df[df.columns.intersection(custom_columns)]
    df = df.reindex(columns=custom_columns)
    df = df.dropna(axis=1, how="all")
    filename = [filename, ".", os.environ["xlformat"]]
    if "csv" in os.environ["xlformat"]:
        df.to_csv("".join(filename), index=False)
    else:
        df.to_excel("".join(filename), index=False)
    if "csv" not in os.environ["xlformat"]:
        wb = Workbook()
        wb = load_workbook("".join(filename))
        ws = wb.worksheets[0]
        for column_cells in ws.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5
        newfilename = os.path.abspath("".join(filename))
        wb.save(newfilename)
    return df


"""
  get criteria details and values
"""


def get_report_details(item, temp, name, criteria):
    if name + "=" in item:
        temp = str(item).split("=", 1)[1]
    return str(temp), criteria


"""
  update figure
"""


def update_fig(fig, type, job, duration):
    fig.update_layout(
        title={"text": "History", "y": 0.97, "x": 0.5,
               "xanchor": "center", "yanchor": "top"},
        xaxis_title=duration,
        yaxis_title="Test Status",
        font=dict(
            family="Trebuchet MS, Helvetica, sans-serif", size=12, color="black",
        ),
        autosize=True,
        hovermode="x unified",
        yaxis={"tickformat": ".0f"},
        xaxis_tickformat="%d/%b/%y",
    )
    fig.update_yaxes(automargin=True)
    if type == "prediction":
        fig.update_layout(
            title={"text": "Trends"}, yaxis_title="Total tests executed", autosize=True,
        )
    return fig


""" 
get styles
"""


def get_style():
    return (
        """
            <style>

                html {{
                height:100%;
                }}
                
                .tabbed {{
                display:  flex;
                text-align: left;
                flex-wrap: wrap;
                box-shadow: 0 0 80px rgba(101, 242, 183, 0.4);
                font-size: 12px;
                font-family: "Trebuchet MS", Helvetica, sans-serif;
                margin:5px;
                width: calc(100% - 10px);
                }}
                .tabbed > input {{
                display: none;
                }}
                .tabbed > input:checked + label {{
                font-size: 14px;
                text-align: center;
                color: white;
                background-image: linear-gradient(to left, #bfee90, #bfee90, black, black,black, #bfee90, #bfee90);
                }}
                .tabbed > input:checked + label + div {{
                color:darkslateblue;
                display: block;
                }}
                .tabbed > label {{
                background-image: linear-gradient(to left, #fffeea,  #333333, #333333 ,#333333 ,#333333 , #333333, #fffeea);
                color: white;
                text-align: center;
                display: block;
                order: 1;
                flex-grow: 1;
                padding: .3%;
                }}
                .tabbed > div {{
                width: calc(100% - 10px);
                order: 2;
                flex-basis: 100%;
                display: none;
                padding: 10px;
                }}

                /* For presentation only */
                .container {{
                width: 100%;
                margin: 0 auto;
                background-color: """
        + os.environ["bgcolor"]
        + """;
                box-shadow: 0 0 20px rgba(400, 99, 228, 0.4);
                }}

                .tabbed {{
                border: 1px solid;
                }}

                hr {{
                background-color: white;
                height: 5px;
                border: 0;
                margin: 10px 0 0;
                }}
                
                hr + * {{
                margin-top: 10px;
                }}
                
                hr + hr {{
                margin: 0 0;
                }}

                .mystyle {{
                    font-size: 12pt;
                    font-family: "Trebuchet MS", Helvetica, sans-serif;
                    border-collapse: collapse;
                    border: 2px solid black;
                    margin:auto;
                    box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
                    background-color: #fffffa;
                }}

                .mystyle body {{
                font-family: "Trebuchet MS", Helvetica, sans-serif;
                    table-layout: auto;
                    position:relative;
                }}

                #slide{{
                width:100%;
                height:auto;
                }}

                #myInput, #myInput2, #myInput3 {{
                background-image: url('http://www.free-icons-download.net/images/mobile-search-icon-94430.png');
                background-position: 2px 4px;
                background-repeat: no-repeat;
                background-size: 25px 30px;
                width: 40%;
                height:auto;
                font-weight: bold;
                font-size: 12px;
                padding: 11px 20px 12px 40px;
                box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
                }}

                p {{
                text-align:center;
                color:white;
                }}

                body {{
                background-color: """
        + os.environ["bgcolor"]
        + """;
                height: 100%;
                background-repeat:  repeat-y;
                background-position: right;
                background-size:  contain;
                background-attachment: initial;
                opacity:.93;
                }}

                h4 {{
                font-family:monospace;
                }}

                @keyframes slide {{
                0% {{
                    transform:translateX(-25%);
                }}
                100% {{
                    transform:translateX(25%);
                }}
                }}

                .mystyle table {{
                    table-layout: auto;
                    width: 100%;
                    height: 100%;
                    position:relative;
                    border-collapse: collapse;
                }}

                tr:hover {{background-color:grey;}}

                .mystyle td {{
                    font-size: 12px;
                    position:relative;
                    padding: 5px;
                    width:10% !important;
                    color: black;
                    border-left: 1px solid #333;
                    border-right: 1px solid #333;
                    background: rgba(255, 253, 207, 0.58);
                    text-align: center;
                    word-break: break-all;
                }}

                table.mystyle td:first-child {{ text-align: left; width:40% !important; }}   

                table.mystyle thead {{
                    background: grey;
                    font-size: 14px;
                    position:relative;
                    border: 1px solid black;
                }}

                table.mystyle thead th {{
                line-height: 200%;
                font-size: 14px;
                font-weight: normal;
                color: #fffffa;
                text-align: center;
                transition:transform 0.25s ease;
                }}

                table.mystyle thead th:hover {{
                    -webkit-transform:scale(1.01);
                    transform:scale(1.01);
                }}

                table.mystyle thead th:first-child {{
                border-left: none;
                }}

                .topnav {{
                overflow: hidden;
                background-color: black;
                opacity: 0.9;
                }}

                .topnav a {{
                float: right;
                display: block;
                color: #333333;
                text-align: center;
                padding: 12px 15px;
                text-decoration: none;
                font-size: 12px;
                position: relative;
                border: 1px solid #6c3;
                font-family: "Trebuchet MS", Helvetica, sans-serif;
                }}

                #summary{{
                box-shadow: 0 0 80px rgba(200, 112, 1120, 0.4);
                position: relative;
                overflow-x: scroll;
                cursor: pointer;
                padding: .1%;
                border-style: outset;
                border-radius: 1px;
                border-width: 1px;
                }}
                
                #logo{{
                box-shadow: 0 0 80px rgba(200, 112, 1120, 0.4);
                position: relative;
                cursor: pointer;
                border-style: outset;
                border-radius: 1px;
                border-width: 1px;
                }}

                .topnav a.active {{
                background-color: #333333;
                color: white;
                font-weight: lighter;
                }}

                .topnav .icon {{
                display: none;
                }}

                @media screen and (max-width: 600px) {{
                .topnav a:not(:first-child) {{display: none;}}
                .topnav a.icon {{
                    color: #DBDB40;
                    float: right;
                    display: block;
                }}
                }}

                @media screen and (max-width: 600px) {{
                .topnav.responsive {{position: relative;}}
                .topnav.responsive .icon {{
                    position: absolute;
                    right: 0;
                    top: 0;
                }}
                .topnav.responsive a {{
                    float: none;
                    display: block;
                    text-align: left;
                }}
                }}

                * {{
                box-sizing: border-box;
                }}

                img {{
                vertical-align: middle;
                }}

                .containers {{
                position: relative;
                }}

                .mySlides {{
                display:none;
                width:90%;
                }}

                #slideshow {{
                cursor: pointer;
                margin:.01% auto;
                position: relative;
                }}

                #ps{{
                height: 10%;
                margin-top: 0%;
                margin-bottom: 90%;
                background-position: center;
                background-repeat: no-repeat;
                background-blend-mode: saturation;
                }}

                #slideshow > div {{
                position: relative;
                width: 90%;
                }}

                #download {{
                background-color: #333333;
                border: none;
                color: white;
                font-size: 12px;
                cursor: pointer;
                }}

                #download:hover {{
                background-color: RoyalBlue;
                }}
                .glow {{
                    font-size: 15px;
                    color: seashell;
                    text-align: center;
                }}
                .reportDiv {{
                    overflow-x: visible;
                    text-align: -webkit-center;
                }}
                .predictionDiv {{
                    overflow-x: auto;
                    text-align: center;
                    display: inline-block;
                    float: left;
                }}
              
                #report{{
                    box-shadow: 0 0 80px rgba(145, 11, 11, 0.4);
                    overflow-x: auto;
                    min-width:70%;
                }}

                #nestle-section{{
                    float:left;
                    width:100%;
                    position:relative;
                }}

                #nestle-section label{{
                    float:left;
                    width:100%;
                    background:#333333;
                    color:rgba(245, 217, 217, 0.99);
                    padding:1px 0;
                    text-align:center;
                    cursor:pointer;
                    border:1px solid #818357;
                }}

                #nestle-section label:hover {{background-color:grey;}}

                #nestle-section .tab-content1{{
                    padding:0 10px;
                    height:0;
                    -moz-transition: height 1s ease;
                    -webkit-transition: height 1s ease;
                    -o-transition: height 1s ease;
                    transition: height 1s ease;
                    overflow:hidden;
                }}

                @media screen and (max-width: 600px) {{
                #nestle-section  input:checked + label + .tab-content1{{
                    padding: 10px;
                    height: auto;
                    -moz-transition: height 1s ease;
                    -webkit-transition: height 1s ease;
                    -o-transition: height 1s ease;
                    transition: height 1s ease;
                    overflow: auto;
                    display: block;
                }}}}
                @media screen and (min-width: 601px) {{
                #nestle-section  input:checked + label + .tab-content1{{
                    padding: 10px;
                    height: auto;
                    -moz-transition: height 1s ease;
                    -webkit-transition: height 1s ease;
                    -o-transition: height 1s ease;
                    transition: height 1s ease;
                    overflow: auto;
                    display: inline;
                    justify-content: center;  
                }}}}

                #nestle-section input:checked + label{{
                    background-color:darkkhaki;
                    color:rgb(15, 61, 16);
                    font-size:16px;
                }}#nestle-section input{{
                    display:none;
                }}
            </style>"""
    )


"""
 get failure html string 
"""


def get_failure_html_string(table):
    bg = os.environ["bgcolor"]
    string = (
        """
        <html lang="en">
       <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
            <link rel="stylesheet" href="https://www.w3schools.com/w3css/4/w3.css">
    		     <head><title aria-label="Report">"""
        + str(os.environ["cloudName"]).upper()
        + """ Failures List</title>
          <meta name="viewport" content="width=device-width, initial-scale=1">
            </head>
        <body style="background-color:"""
        + bg
        + """;">
        <style>
        .reportDiv {
                    overflow-x: visible;
                    text-align: center;
        }
        .mystyle {
            font-size: 12pt;
            font-family: "Trebuchet MS", Helvetica, sans-serif;
            border-collapse: collapse;
            border: 2px solid black;
            margin: auto;
            box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
            background-color: #fffffa;
            margin: 3%;
        }
        
        .mystyle body {
            font-family: "Trebuchet MS", Helvetica, sans-serif;
            table-layout: auto;
            position: relative;
        }
        
        .mystyle table {
            table-layout: auto;
            width: 100%;
            height: 100%;
            position: relative;
            border-collapse: collapse;
        }
        
        tr:hover {
            background-color: rgba(190, 240, 196, 0.863);
        }
        
        .mystyle td {
            box-sizing: border-box;
            font-size: 12px;
            position: relative;
            padding: 5px;
            width: 10%;
            color: black;
            border-left: 1px solid #333;
            border-right: 1px solid #333;
            background: rgba(255, 253, 207, 0.58);
        }
        
        .mystyle tr {
            text-align: center;
        }
        .mystyle thead {
            box-sizing: border-box;
            background: tan;
            color: black;
            font-size: 14px;
            position: relative;
            border: 1px solid black;
        }
        
        .mystyle th {
            box-sizing: border-box;
            line-height: 200%;
            font-size: 14px;
            background: tan;
            font-weight: bold;
            color: black;
            text-align: center;
            transition: transform 0.25s ease;
        }
        
        table.mystyle>tbody>tr>td:nth-of-type(1) {
            width: .5%;
            text-align: center;
        }
        
        table.mystyle>tbody>tr>td:nth-of-type(2) {
            width: 5%;
            text-align: left;
        }
        
        table.mystyle>tbody>tr>td:nth-of-type(3) {
            width: 10%;
            text-align: left;
        }
        table.mystyle>tbody>tr>td:nth-of-type(5) {
            width: 1%;
        }
        table.mystyle>tbody>tr>td:nth-of-type(6) {
            width: .5%;
        }
        table.mystyle>tbody>tr>td:nth-of-type(7) {
            width: .6%;
        }
        #itable {
            min-width: 60%;
            max-width: 95%;
            word-break: break-all;
        }
    </style>
        <script src="https://ajax.googleapis.com/ajax/libs/jquery/3.4.1/jquery.min.js"></script>
            <script>
                $(document).ready(function() {
                    var table = document.getElementById("itable");
                    var rowCount = table.rows.length;
                    for (var i = 0; i < rowCount; i++) {
                        if (i >= 1) {
                            tcNameColumn = 1;
                            reportLink = 3;
                            var txt = table.rows[i].cells[tcNameColumn].innerHTML;
                            var url = table.rows[i].cells[reportLink].innerHTML;
                            var row = $('<tr></tr>')
                            var link = document.createElement("a");
                            link.href = url;
                            link.innerHTML = txt;
                            link.target = "_blank";
                            table.rows[i].cells[tcNameColumn].innerHTML = '';
                            table.rows[i].cells[tcNameColumn].appendChild(link);
                        }
                    }
                    $("table").find("tbody th").hide();
                    $("#itable").find("th, td").filter(":nth-child(" + (reportLink + 1) + ")").hide();
                    $("#search").on("keyup", debounce(function() {
                        var value = $(this).val().toLowerCase();
                        $("#itable tbody tr").filter(function() {
                            $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
                        });
                    }, 300));

                    function debounce(func, wait, immediate) {
                        var timeout;
                        return function() {
                            var context = this,
                                args = arguments;
                            var later = function() {
                                timeout = null;
                                if (!immediate) func.apply(context, args);
                            };
                            var callNow = immediate && !timeout;
                            clearTimeout(timeout);
                            timeout = setTimeout(later, wait);
                            if (callNow) func.apply(context, args);
                        };
                    };
                });
            </script>
            <meta name="viewport" content="width=device-width, initial-scale=1">
        </head>
        <body>
        <div class="reportDiv"><br>
        <input id="search" aria-label="search" type="text" placeholder="Search/Filter..">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</p>
        """ + table + """ </div></body></html>""")
    return str(string)


"""
 get html string
"""


def get_html_string(graphs, tagrec, tagLinks, failurereasons, monthlyStats, topfailedtable, execution_summary):
    bg = os.environ["bgcolor"]
    heading = os.environ["title"]
    string = (
        """
    <html lang="en">
       <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
            <link rel="stylesheet" href="https://www.w3schools.com/w3css/4/w3.css">
    		     <head><title aria-label="Report">"""
        + heading
        + """</title>
          <meta name="viewport" content="width=device-width, initial-scale=1">
            </head>
    <body style="background-color:"""
        + bg
        + """;">
    """
        + get_style()
        + """ <div style="
          overflow-x: auto;
          text-align: center;
          color: #e5e7cc;
          background-color: #22283a;
          font-family: Helvetica, Arial, sans-serif;
          font-size: 24px;
          font-weight: bold;
          padding: 5px 5px 10px 10px;
          box-shadow: 0 4px 8px 0 rgba(0, 0, 0, 0.2), 0 6px 20px 0 rgba(0, 0, 0, 0.19);
        ">
        <a href="https://"""
        + str(os.environ["cloudName"])
        + """.perfectomobile.com" target="_blank" class="site-logo">
                            <img id="logo" style="height:30px !important;"  src="""
        + os.environ["company_logo"]
        + """ style="margin:1%;" alt="Company logo" ></a>
        """ + heading + """</div>"""
        + """<div style="padding:1px 0; background-color:rgba(236, 213, 171, 0.56); color:black;text-align:center;font-family:Verdana,sans-serif;font-size:16px;width:100%;cursor:pointerfont-weight:bold">
        """
        + title +
        """
        </div><div id="nestle-section">
        <input type="radio" id="tab1" name="tabs1" checked=""/><label for="tab1">Summary Report</label>
        <div class="tab-content1"> </p><div class="reportDiv"> """
        + execution_summary
        + """ alt='execution summary' id='reportDiv'> </img></br></div></div>"""
        + """<div class="reportDiv">"""
        + execution_status
        + per_job_status
        + per_tag_status
        + """ </div></br><input type="radio" id="tab2" name="tabs" checked=""/><label for="tab2">OS Summary</label><div class="tab-content1">
          <div class="reportDiv">"""
        + monthlyStats
        + issues
        + """ <br><a href="./""" +
        str(os.environ["cloudName"]) + """_unique_failures.html" style="white-space:nowrap;text-decoration:none;background-color: rgb(195, 132, 49);color: white;padding: 8px 16px;font-family: sans-serif;border-radius: 3px;">All Unique Failures</a>"""
        + tagLinks
        + """</p></div></div><input type="radio" id="tab4" name="tabs" checked=""/><label for="tab4">Custom Failure Reasons</label><div class="tab-content1">
          <div class="reportDiv">"""
        + failurereasons
        + """ </div></div><input type="radio" id="tab5" name="tabs" checked=""/><label for="tab5">Top Failed Tests</label><div class="tab-content1">
          <div class="reportDiv">"""
        + topfailedtable
        + """ </div>
          </div><input type="radio" id="tab6" name="tabs" checked=""/><label for="tab6">Top Recommendations</label><div class="tab-content1">
          <div class="reportDiv">"""
        + recommendations + tagrec
        + """ </div></div><div class="reportDiv">"""
        + "".join(graphs)
        + """</div></body>"""
    )
    return str(string)


"""
 get html string email
"""


def get_html_string_email(graphs, tagrec, failurereasons, monthlyStats, topfailedtable, execution_summary):
    bg = os.environ["bgcolor"]
    heading = os.environ["title"]
    header = 'style="box-sizing: border-box; float: left; width: 100%; padding: 1px 0; text-align: center; cursor: pointer; font-size: 16px; color: black; background-color: darkkhaki; border: 3px solid antiquewhite;"'
    tabcontent = (
        'style="box-sizing: border-box; padding: 10px; height: auto; -moz-transition: height 1s ease; -webkit-transition: height 1s ease; -o-transition: height 1s ease; transition: height 1s ease; overflow: auto; display: block;background-color:'
        + bg
        + ';'
    )
    reportDiv = 'style="box-sizing: border-box; overflow-x: auto; text-align: center;background-color:' + \
        os.environ["bgcolor"]+'"'

    string = (
        """
    <html lang="en">
       <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
            <link rel="stylesheet" href="https://www.w3schools.com/w3css/4/w3.css">
    		     <head><title aria-label="Report">"""
        + heading
        + """</title><meta name="viewport" content="width=device-width, initial-scale=1">
            </head>
            <style> table.mystyle td:first-child {{ text-align: left; width:40% !important; }}</style>
    <body style="box-sizing: border-box; height: 100%; background-repeat: repeat-y; background-position: right; background-size: contain; background-attachment: initial; opacity: .93; background-color:"""
        + bg
        + """;"> <div style="
          overflow-x: auto;
          text-align: center;
          color: #e5e7cc;
          background-color: #22283a;
          font-family: Helvetica, Arial, sans-serif;
          font-size: 24px;
          font-weight: bold;
          padding: 5px 5px 10px 10px;
          box-shadow: 0 4px 8px 0 rgba(0, 0, 0, 0.2), 0 6px 20px 0 rgba(0, 0, 0, 0.19);
        ">
        <a href="https://"""
        + str(os.environ["cloudName"])
        + """.perfectomobile.com" target="_blank" class="site-logo">
                            <img id="logo" style="height:30px !important;"  src="""
        + os.environ["company_logo"]
        + """ style="margin:1%;" alt="Company logo" ></a>
        """ + heading + """</div><div style="padding:1px 0; background-color:rgba(236, 213, 171, 0.56); color:black;text-align:center;font-family:Verdana,sans-serif;font-size:16px;width:100%;cursor:pointerfont-weight:bold">
        """
        + title +
        """<div """
        + header
        + """><b><center>Summary Report</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + """">
        <div class="reportDiv" """
        + reportDiv
        + """> """
        + execution_summary
        + """ alt='execution summary' id='reportDiv' """
        + reportDiv
        + """> </img></br></div></div><div class="reportDiv" """
        + reportDiv
        + """>"""
        + execution_status
        + per_job_status_email
        + per_tag_status_email
        + """<br></div><div """
        + header
        + """</div>"""
        + """<b><center>OS Summary</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + 'display: inline-block !important;"'
        + """>
        <div class="reportDiv" """
        + reportDiv
        + """>"""
        + monthlyStats
        + issues_email
        + """<br></div></div><div """
        + header
        + """><b><center>Custom Failure Reasons</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + """>
        <div class="reportDiv" """
        + reportDiv
        + """>"""
        + failurereasons
        + """<br></div></div><div """
        + header
        + """><b><center>Top Failed Tests</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + """>
        <div class="reportDiv" """
        + reportDiv
        + """>"""
        + topfailedtable
        + """<br></div></div><div """
        + header
        + """><b><center>Top Recommendations</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + """>
        <div class="reportDiv" """
        + reportDiv
        + """>"""
        + recommendations + tagrec
        + """ </div></div></div><br><div class="reportDiv">"""
        + "".join(graphs)
        + """</div><br></body>"""
    )
    return str(string)


def style_df_email(df):
    return (
        df.replace(
            '<table border="1" class="dataframe mystyle" id="report">',
            '<table border="1" class="dataframe mystyle" id="report" style="box-sizing: border-box; font-size: 12pt; font-family:Trebuchet MS, Helvetica, sans-serif; border-collapse: collapse; border: 2px solid black;  margin-left:5%;margin-right:5%; margin:auto;background-color: #fffffa; box-shadow: 0 0 30px rgba(145, 11, 11, 0.4); overflow-x: auto; min-width: 60%; max-width:80%;" bgcolor="#fffffa">',
        )
        .replace("<tr>", '<tr style="box-sizing: border-box;" align="center">')
        .replace(
            "<thead>",
            '<thead style="box-sizing: border-box; background: tan; color:black;font-size: 14px; position: relative; border: 1px solid black;">',
        )
        .replace(
            "<th>",
            '<th style="box-sizing: border-box; line-height: 200%; font-size: 14px; background: tan; font-weight: bold; color: black; text-align: center; transition: transform 0.25s ease;" align="center">',
        )
        .replace("<tbody>", '<tbody style="box-sizing: border-box;">')
        .replace(
            "<td>",
            '<td style="box-sizing: border-box; font-size: 12px; position: relative; padding: 5px; color: black; border-left: 1px solid #333; border-right: 1px solid #333; background: rgba(255, 253, 207, 0.58);width:15%;" align="center">',
        )
    )


def createHref(jobName, tag, status, content):
    jobList = ""
    tagList = ""
    jobNumList = ""
    if(isinstance(content, int) and int(content) == 0):
        return content
    else:
        if(startDate):
            if "-" not in startDate:
                startExecutionTime = startDate
            else:
                startExecutionTime = pastDateToMS(startDate, 0)
            if "-" not in endDate:
                endExecutionTime = endDate
            else:
                endExecutionTime = round(int(datetime.strptime(
                    str(endDate) + " 23:59:59,999", "%Y-%m-%d %H:%M:%S,%f"
                ).timestamp() * 1000))
            url = "https://" + os.environ["cloudName"] + ".app.perfectomobile.com/reporting/library?startExecutionTime[0]=" + str(
                startExecutionTime) + "&startExecutionTime[1]=custom&endExecutionTime[0]=" + str(endExecutionTime)
        else:
            url = "https://" + os.environ["cloudName"] + \
                ".app.perfectomobile.com/reporting/library?"
        if(jobName and jobName != "All Jobs"):
            for i, job in enumerate(jobName.split(";")):
                jobList = jobList + "&jobName[" + str(i) + "]=" + job
            url = url + jobList
        if(jobNumber):
            for i, jobNum in enumerate(jobNumber.split(";")):
                jobNumList = jobNumList + \
                    "&jobNumber[" + str(i) + "]=" + jobNum
            url = url + jobNumList
        if(tag):
            for i, job in enumerate(tag.split(";")):
                tagList = tagList + "&tags[" + str(i) + "]=" + job
            url = url + tagList
        if(status):
            url = url + '&status[0]='+status+'&_search=&_searchViewAll=false'
        else:
            url = url + '&_search=&_searchViewAll=false'

        url = '<a target="_blank" href="' + \
            str(url) + '">' + str(content) + "</a>"
        return url


def get_recommendations(df, failed_blocked, topfailedTCNames, failed, passed, blocked, name, bool):
    labIssuesCount = 0
    totalFailCount = 0
    totalPassCount = 0
    totalUnknownCount = 0
    totalTCCount = 0
    scriptingIssuesCount = 0
    appCrashIssuesCount = 0
    testDataIssuesCount = 0
    environmentIssuesCount = 0
    orchestrationIssuesCount = 0
    cleanedFailureList = {}
    # recommendations
    totalFailCount = failed.shape[0]
    totalPassCount = passed.shape[0]
    blockedCount = blocked.shape[0]
    # failures count
    failuresmessage = []
    if debug == "true":
        failureListFileName = name + "_failures" + '.txt'
        print("transfering all failure reasons to: %s" %
              (os.path.join(os.path.abspath(os.curdir), failureListFileName)))
        open(failureListFileName, 'w').close
    totalUnknownCount = df[(df["Test Status"] == "UNKNOWN")].shape[0]
    totalTCCount = df.shape[0]
    failed_blocked = failed_blocked[~failed_blocked.name.isin(["Interactive session"])]
    if len(failed_blocked) > 0:
        failuresmessage = (
            failed_blocked.groupby(["message"])
            .size()
            .reset_index(name="#Failed")
            .sort_values("#Failed", ascending=False)
        )
        # Get all errors and links
        i = 1
        allFailuresResultLink = []
        for error, count in failuresmessage.itertuples(index=False):
            report_link = df.loc[df['message'].str.startswith(
                str(error.strip()), na=False), "reportURL"].iloc[0]
            testName = df.loc[df['reportURL'].str.startswith(
                str(report_link), na=False), "name"].iloc[0]
            job = str(df.loc[df['reportURL'].str.startswith(
                str(report_link), na=False), "job/name"].iloc[0]).replace('nan', '')
            try:
                jobNumber = str(int(str(int(df.loc[df['reportURL'].str.startswith(
                    str(report_link), na=False), "job/number"].iloc[0])).replace('nan', '')))
            except ValueError:
                jobNumber = ''
            owner = str(df.loc[df['reportURL'].str.startswith(
                str(report_link), na=False), "owner"].iloc[0])
            miniJson = {
                "#": i,
                "Test Name": testName,
                "Error": error,
                "link": report_link,
                "Job Name": job,
                "Job #": jobNumber,
                "Owner": owner
            }
            allFailuresResultLink.append(miniJson)
            i += 1
        allFailuresResultLink = json.dumps(allFailuresResultLink)
        output = json2html.convert(json=allFailuresResultLink)
        file = open(name + '_unique_failures.html', 'w')
        file.write(get_failure_html_string(str(output).replace(
            '<table border="1">', '<table class="dataframe mystyle" border="1" id="itable">')))
        file.close()

        global labIssues
        global orchestrationIssues
        for commonError, commonErrorCount in failuresmessage.itertuples(
            index=False
        ):
            for labIssue in labIssues:
                if re.search(labIssue, commonError):
                    labIssuesCount += commonErrorCount
                    break
            for orchestrationIssue in orchestrationIssues:
                if re.search(orchestrationIssue, commonError):
                    orchestrationIssuesCount += commonErrorCount
                    break
            error = str(commonError)
            regex = ""
            if os.environ["regex"] != "":
                regex = "|" + os.environ["regex"]
            regEx_Filter = "Build info:|For documentation on this error|at org.xframium.page|Scenario Steps:| at WebDriverError|\(Session info:|XCTestOutputBarrier\d+|\s\tat [A-Za-z]+.[A-Za-z]+.|View Hierarchy:|Got: |Stack Trace:|Report Link|at dalvik.system|Output:\nUsage|t.*Requesting snapshot of accessibility|\{ Error\:|at\sendReadableNT|at\sFunction|\sat\smakeRequest|at\sObject\.\_errnoException|\"stack\"\:|('|)\n.*Error\:\s|at\sRequest.callback|\n\s+at\s" + regex
            if re.search(regEx_Filter, error, re.MULTILINE):
                if re.search("('|)(\n|)AssertionError\:.*\n\s+at.*\n.*at", error, re.MULTILINE):
                    error = "at".join(error.split("at", 2)[
                                      :2]).replace(r'\n', '\n')

                else:
                    error = str(re.compile(regEx_Filter).split(error)[0])
                    if "An error occurred. Stack Trace:" in error:
                        error = error.split(
                            "An error occurred. Stack Trace:")[1]
            if re.search("error: \-\[|Fatal error:", error):
                error = str(
                    re.compile("error: \-\[|Fatal error:").split(error)[1]
                )
            if re.search("\n[*]{4}", error):
                error = str(
                    re.compile("\n[*]{4}").split(error)[1]
                )
                error = str(
                    re.compile("[*]{4}\n").split(error)[0]
                )
            if re.search("ERROR: com.", error):
                error = str(
                re.compile("ERROR:").split(error)[1]
            )
            error = str(
                re.compile("\n").split(error)[0]
            )
            if error.strip() in cleanedFailureList:
                cleanedFailureList[error.strip()] += 1
            else:
                cleanedFailureList[error.strip()] = commonErrorCount
            appCrashIssuesCount = len(
                df.loc[df['Custom Failure Reason'] == "Application crashed"])
            environmentIssuesCount = len(
                df.loc[df['message'].str.contains("Error: Request failed with", na=False)])
            testDataIssuesCount = len(
                df.loc[df['message'].str.contains("TEST_DATA_ERROR", na=False)])
            scriptingIssuesCount = (totalFailCount + blockedCount) - (
                orchestrationIssuesCount + labIssuesCount + appCrashIssuesCount +
                environmentIssuesCount + testDataIssuesCount
            )
            if debug == "true":
                with open(failureListFileName, "a", encoding="utf-8") as myfile:
                    myfile.write(str(error.strip()) +
                                 '\n*******************************************\n')

    # Top 5 failure reasons
    topFailureDict = {}
    failureDict = Counter(cleanedFailureList)
    count_total = int(str(os.environ["recommendations"]))
    for commonError, commonErrorCount in failureDict.most_common(count_total):
        topFailureDict[commonError] = int(commonErrorCount)
    suggesstionsDict = {}
    # reach top errors and clean them
    i = 0
    for commonError, commonErrorCount in topFailureDict.items():
        if "ERROR: No device was found" in commonError:
            error = (
                "Raise a support case for the error: *|*"
                + commonError.strip() + "*|*"
            )
        elif "Cannot open device" in commonError:
            error = (
                "Reserve the device/ use perfecto lab auto selection feature to avoid the error:  *|*"
                + commonError.strip() + "*|*"
            )
        elif (
            '(UnknownError) Failed to execute command button-text click: Needle not found for expected value: "Allow" (java.lang.RuntimeException)'
            in commonError
        ):
            error = (
                "Allow text/popup was not displayed as expected. It could be an environment issue as the error: *|*"
                + commonError.strip() + "*|*"
            )
        else:
            error = (normalize("NFKD", str(commonError.strip())))
        report_link = df.loc[df['message'].str.contains(
            re.escape(str(commonError.strip())), na=False), "reportURL"].iloc[0]
        suggesstionsDict[error] = [int(commonErrorCount), report_link]
    eDict = {}
    eDict = edict(
        {
            "status": [
                {
                    "#Total": "Count ->",
                    "#Executions": createHref(jobName, reportTag, '', totalTCCount),
                    "#Pass": createHref(jobName, reportTag, 'PASSED', totalPassCount),
                    "#Failed": createHref(jobName, reportTag, 'FAILED', totalFailCount),
                    "#Blocked": createHref(jobName, reportTag, 'BLOCKED', blockedCount),
                    "#Unknowns": createHref(jobName, reportTag, 'UNKNOWN', totalUnknownCount),
                    "Overall Pass %": str(
                        int(percentageCalculator(totalPassCount, totalTCCount))
                    )
                    + "%",
                },
            ],
            "issues": [
                {
                    "#Issues": "Count ->",
                    "#Scripting": scriptingIssuesCount,
                    "#App Crash": appCrashIssuesCount,
                    "#Environment Issues": environmentIssuesCount,
                    "#Test Data Issues": testDataIssuesCount,
                    "#Lab": labIssuesCount,
                    "#Orchestration": orchestrationIssuesCount,
                },
            ],
            "recommendation": [
            ],
        }
    )
    header = 'style="box-sizing: border-box; float: left; width: 100%; padding: 1px 0; text-align: center; cursor: pointer; font-size: 16px; color: black; background-color: darkkhaki; border: 3px solid antiquewhite;"'
    perJob = []
    tag = ''
    for i, job in enumerate(jobName.split(";")):
        if(job != ""):
            job_df = df.loc[(df["job/name"] == job)]
            job_total = job_df.shape[0]
            failed = job_df[(job_df["Test Status"] == "FAILED")].shape[0]
            passed = job_df[(job_df["Test Status"] == "PASSED")].shape[0]
            blocked = job_df[(job_df["Test Status"] == "BLOCKED")].shape[0]
            unknown = job_df[(job_df["Test Status"] == "UNKNOWN")].shape[0]
            perJob.append({"Job": createHref(job, "", '', job), "#Executions": createHref(job, "", '', job_total), "#Pass": createHref(job, "", 'PASSED', passed), "#Failed": createHref(
                job, "", 'FAILED', failed), "#Blocked": createHref(job, "", 'BLOCKED', blocked), "#Unknowns": createHref(job, "", 'UNKNOWN', unknown), "Overall Pass %": str(int(percentageCalculator(passed, job_total))) + "%"})

    global per_job_status
    global per_job_status_email
    per_job_status = ""
    per_job_status_email = ""
    per_job = DataFrame.from_dict(perJob)

    if(per_job.shape[0] > 0):
        per_job = per_job.to_html(
            classes="mystyle",
            table_id="report",
            index=False,
            render_links=True,
            escape=False,
        )
        per_job_status = """ <br><label for="tab2" style="background:  darkkhaki !important; color:rgb(15, 61, 16) !important;">Job wise Summary</label><br><br>""" + style_df_email(
            per_job)
        per_job_status_email = style_df_email(per_job)
        per_job_status_email = """ <br><div """ + header + \
            """><center><label for="tab2" style="background:  darkkhaki !important; color:black !important;font-weight:bold;">Job wise Summary</label></center></div><br><br>""" + per_job_status_email

    perTag = []
    tags_cols = [col for col in df.columns if 'tags' in col]
    exp_tags = ""
    if(reportTag == ''):
        exp_tags = os.environ["recommend_tag"]
    else:
        exp_tags = reportTag
    for i, tag in enumerate(exp_tags.split(";")):
        if(tag != ""):
            query = ''
            for i, col in enumerate(tags_cols):
                query += '`' + col + "` == '" + tag + "'"
                if(i != (len(tags_cols) - 1)):
                    query += ' or '
            tag_df = df.query(query)
            job_total = tag_df.shape[0]
            failed = tag_df[(tag_df["Test Status"] == "FAILED")].shape[0]
            passed = tag_df[(tag_df["Test Status"] == "PASSED")].shape[0]
            blocked = tag_df[(tag_df["Test Status"] == "BLOCKED")].shape[0]
            unknown = tag_df[(tag_df["Test Status"] == "UNKNOWN")].shape[0]
            perTag.append({"Tag": createHref(jobName, tag, '', tag), "#Executions": createHref(jobName, tag, '', job_total), "#Pass": createHref(jobName, tag, 'PASSED', passed), "#Failed": createHref(
                jobName, tag, 'FAILED', failed), "#Blocked": createHref(jobName, tag, 'BLOCKED', blocked), "#Unknowns": createHref(jobName, tag, 'UNKNOWN', unknown), "Overall Pass %": str(int(percentageCalculator(passed, job_total))) + "%"})

    global per_tag_status
    global per_tag_status_email
    per_tag_status = ""
    per_tag_status_email = ""
    per_tag = DataFrame.from_dict(perTag)

    if(per_tag.shape[0] > 0):
        per_tag = per_tag.to_html(
            classes="mystyle",
            table_id="report",
            index=False,
            render_links=True,
            escape=False,
        )
        per_tag_status = """ <br><label for="tab2" style="background:  darkkhaki !important; color:rgb(15, 61, 16) !important;">Tag wise Summary</label><br><br>""" + style_df_email(
            per_tag)
        per_tag_status_email = style_df_email(per_tag)
        per_tag_status_email = """ <br><div """ + header + \
            """><center><label for="tab2" style="background:  darkkhaki !important; color:black !important;font-weight:bold;">Tag wise Summary</label></center></div><br><br>""" + per_tag_status_email

    recommendations_count = int(str(os.environ["recommendations"]))
    i = 0
    dynamic_rec = []
    while i < recommendations_count:
        dynamic_rec.append({"Recommendations": "-", "Occurences": "-",
                           "ReportURL": "-", "Rank": i+1, "impact": "0"})
        i += 1
    eDict['recommendation'] = dynamic_rec
    jsonObj = edict(eDict)

    if(bool):
        if float(percentageCalculator(totalUnknownCount, totalTCCount)) >= 30:
            suggesstionsDict[
                "# Fix the unknowns. The unknown script ratio is too high (%) : "
                + str(percentageCalculator(totalUnknownCount, totalTCCount))
                + "%"
            ] = [percentageCalculator(
                totalPassCount + totalUnknownCount, totalTCCount
            ) - percentageCalculator(
                totalPassCount, totalTCCount
            ), "-"]
        if len(suggesstionsDict) < count_total:
            if (topfailedTCNames.shape[0]) > 1:
                for tcName, status in topfailedTCNames.itertuples(index=False):
                    suggesstionsDict[
                        "# Fix the top failing tests listed under 'Top Failed Tests' "
                    ] = [1, "-"]
                    break

        if len(suggesstionsDict) < count_total:
            if int(percentageCalculator(totalFailCount, totalTCCount)) > 15:
                if totalTCCount > 0:
                    suggesstionsDict[
                        "# Fix the failures. The total failures % is too high (%) : "
                        + str(percentageCalculator(totalFailCount, totalTCCount))
                        + "%"
                    ] = [int(percentageCalculator(totalFailCount, totalTCCount)), "-"]
        if len(suggesstionsDict) < count_total:
            if float(percentageCalculator(totalPassCount, totalTCCount)) < 80 and (
                totalTCCount > 0
            ):
                suggesstionsDict[
                    "# Fix the failures. The total pass %  is too less (%) : "
                    + str(int(percentageCalculator(totalPassCount, totalTCCount)))
                    + "%"
                ] = [int(
                    (100
                     - (
                         percentageCalculator(
                             totalPassCount + totalUnknownCount, totalTCCount
                         )
                         - percentageCalculator(totalPassCount, totalTCCount)
                     )
                     ) - int(percentageCalculator(totalPassCount, totalTCCount))), "-"]
    if len(suggesstionsDict) < count_total:
        if totalTCCount == 0:
            suggesstionsDict[
                "# There are no executions for today. Try Continuous Integration with any tools like Jenkins and schedule your jobs today. Please reach out to Professional Services team of Perfecto for any assistance :) !"
            ] = [100, "-"]
        elif int(percentageCalculator(totalPassCount, totalTCCount)) > 80:
            print(str(int(percentageCalculator(totalPassCount, totalTCCount))))
            suggesstionsDict["# Great automation progress. Keep it up!"] = [
                0, "-"]

        int(percentageCalculator(totalFailCount, totalTCCount)) > 15
    counter = 0
    totalImpact = 0
    for sugg, commonErrorCount in sorted(suggesstionsDict.items(), key=lambda x: (x[1], x[0]), reverse=True)[:count_total]:
        impact = 1
        if str(sugg) != "" and str(commonErrorCount[0]) != "":
            if sugg.startswith("# "):
                jsonObj.recommendation[counter].ReportURL = '-'
                sugg = sugg.replace("# ", "")
                impact = str(int(float(str(commonErrorCount[0]))))
                jsonObj.recommendation[counter].Occurences = "-"
            else:
                jsonObj.recommendation[counter].Occurences = int(
                    float(str(commonErrorCount[0])))
                jsonObj.recommendation[counter].ReportURL = '<a target="_blank" href="' + \
                    commonErrorCount[1] + '">link</a>'
                impact = str(percentageCalculator(
                    totalPassCount +
                    int(float(str(commonErrorCount[0]))), totalTCCount
                ) - percentageCalculator(totalPassCount, totalTCCount))
            jsonObj.recommendation[counter].impact = (
                str(("%.2f" % round(int(float(str(impact))), 2))) + "%"
            )
            jsonObj.recommendation[counter].Recommendations = escape(
                sugg.replace("*|*", "'")
                .replace("{", "{{")
                .replace("}", "}}")
                .strip()
            )
            totalImpact += round(int(float(str(impact))), 2)
        counter += 1
    global execution_status
    execution_status = DataFrame.from_dict(jsonObj.status)
    execution_status = execution_status.to_html(
        classes="mystyle",
        table_id="report",
        index=False,
        render_links=True,
        escape=False,
    )
    execution_status = style_df_email(execution_status)
    global issues
    issues = ""
    global issues_email
    issues_email = ""
    if("true" in str(os.environ["showIssues"])):
        issues_df = DataFrame.from_dict(jsonObj.issues)
        issues_ori = issues_df.to_html(
            classes="mystyle",
            table_id="report",
            index=False,
            render_links=True,
            escape=False,
        )
        header = 'style="box-sizing: border-box; float: left; width: 100%; padding: 1px 0; text-align: center; cursor: pointer; font-size: 16px; color: black; background-color: darkkhaki; border: 3px solid antiquewhite;"'
        issues = """</div></div><input type="radio" id="tab3" name="tabs" checked=""/><label for="tab3">Issues</label><div class="tab-content1">
          <div class="reportDiv">""" + issues_ori
        issues = style_df_email(issues)
        tabcontent = 'style="box-sizing: border-box; padding: 10px; height: auto; -moz-transition: height 1s ease; -webkit-transition: height 1s ease; -o-transition: height 1s ease; transition: height 1s ease; overflow: auto; display: inline;justify-content: center;"'
        reportDiv = (
            'style="box-sizing: border-box; overflow-x: visible; text-align: -webkit-center;"'
        )
        issues_email = str("""<br></div><div """
                           + header
                           + """</div>"""
                           + """<b><center>Issues</center></b></label></div><div class="tab-content1" """
                           + tabcontent
                           + """>
        <div class="reportDiv" """
                           + reportDiv
                           + """>""") + style_df_email(issues_ori)
    else:
        issues = """</div></div><input type="radio" id="tab3" name="tabs" checked=""/><label for="tab3">Groups</label><div class="tab-content1">
          <div class="reportDiv">""" + issues
    recommendations = DataFrame.from_dict(jsonObj.recommendation)
    # recommendations = recommendations.sort_values(['impact'], ascending=False)
    if totalImpact > 100:
        recommendations.columns = [
            "Issues",
            "Occurences",
            "ReportURL",
            "Rank",
            "Pass% Increase",
        ]
    else:
        recommendations.columns = [
            "Issues",
            "Occurences",
            "ReportURL",
            "Rank",
            "Pass% Increase - " + str(round(totalImpact, 2)) + "%",
        ]
    recommendations = recommendations[
        recommendations.Issues.astype(str) != "-"
    ]
    recommendations = recommendations.to_html(
        classes="mystyle",
        table_id="report",
        index=False,
        render_links=True,
        escape=False,
    )
    return style_df_email(recommendations)


def process_failures(df, failed_blocked, topfailedTCNames, graphs, interactive_graphs, failed, passed, blocked, failurereasons, monthlyStats, topfailedtable, execution_summary):

    global recommendations
    recommendations = get_recommendations(
        df, failed_blocked, topfailedTCNames, failed, passed, blocked, str(os.environ["cloudName"]), True)

    tags_cols = [col for col in df.columns if 'tags' in col]
    recommend_tag = os.environ["recommend_tag"]
    tagrec = ""
    tagLinks = ""
    if recommend_tag != "":
        for i, tag in enumerate(recommend_tag.split(";")):
            query = ''
            for i, col in enumerate(tags_cols):
                query += '`' + col + "` == '" + tag + "'"
                if(i != (len(tags_cols) - 1)):
                    query += ' or '
            tag_df = df.query(query)
            tagrec += """ <br><div style="padding:1px 0; background-color:darkkhaki; color:black;text-align:center;font-family:Verdana,sans-serif;font-size:16px;width:100%;cursor:pointer;">
                        """ + tag + """ Recommendations
                </div>"""
            tagrec += get_recommendations(df, tag_df, topfailedTCNames,
                                          failed, passed, blocked, str(tag), False)
            tagLinks += """ &nbsp; <a href="./""" + \
                str(tag) + """_unique_failures.html" style="white-space:nowrap;text-decoration:none;background-color: rgb(195, 132, 49);color: white;padding: 8px 16px;font-family: sans-serif;border-radius: 3px;">""" + str(tag) + """ Failures</a>"""

    # prepares graphs & interactive graphs
    with open(email_report_filename, "a") as f:
        f.write(
            get_html_string_email(graphs, tagrec, failurereasons, monthlyStats, topfailedtable, execution_summary).format(
                table=df.to_html(
                    classes="mystyle",
                    table_id="report",
                    index=False,
                    render_links=True,
                    escape=False,
                )
            )
        )
    graphs.clear()
    with open(live_report_filename, "a") as f:
        f.write(
            get_html_string(graphs, tagrec, tagLinks, failurereasons, monthlyStats, topfailedtable, execution_summary).format(
                table=df.to_html(
                    classes="mystyle",
                    table_id="report",
                    index=False,
                    render_links=True,
                    escape=False,
                )
            )
        )
    with open(live_report_filename, "a") as f:
        f.write(''.join(interactive_graphs)
                )


def main():
    """
    Runs the perfecto actions and reports
    """
    try:
        start_time = datetime.now().replace(microsecond=0)
        freeze_support()
        init()
        #     """fix Python SSL CERTIFICATE_VERIFY_FAILED"""
        if not os.environ.get("PYTHONHTTPSVERIFY", "") and getattr(
            ssl, "_create_unverified_context", None
        ):
            ssl._create_default_https_context = ssl._create_unverified_context
        parser = ArgumentParser(
            description="Perfecto Actions Reporter")
        parser.add_argument(
            "-c",
            "--cloud_name",
            metavar="cloud_name",
            help="Perfecto cloud name. (E.g. demo) or add it as a cloudName environment variable",
            nargs="?",
        )
        parser.add_argument(
            "-s",
            "--security_token",
            metavar="security_token",
            type=str,
            help="Perfecto Security Token/ Pass your Perfecto's username and password in user:password format  or add it as a securityToken environment variable",
            nargs="?",
        )
        parser.add_argument(
            "-o",
            "--output",
            type=str,
            metavar="output in html",
            help="output in html. Values: true/false. Default is true",
            nargs="?",
        )
        parser.add_argument(
            "-l",
            "--logo",
            type=str,
            metavar="shows customer logo",
            help="shows client logo if valid official client website url is specified in this sample format: www.perfecto.io",
            nargs="?",
        )
        parser.add_argument(
            "-e",
            "--email",
            type=str,
            metavar="prepares AI based emailable and live report along with statistics & recommendations",
            help="creates a downloadable csv/xlsx of reporting data along with AI emailable & live report with live charts, AI predictions and recommendations.",
            nargs="?",
        )
        parser.add_argument(
            "-b",
            "--bgcolor",
            type=str,
            metavar="sets the background color in report",
            help="overrides the background color in report based on provided hex color",
            nargs="?",
        )
        parser.add_argument(
            "-d",
            "--debug",
            type=str,
            metavar="Shows more logs and output txt files",
            help="Shows more logs and creates txt file of all failures as well as tag failures",
            nargs="?",
        )
        args = vars(parser.parse_args())
        try:
            if not args["cloud_name"]:
                print("Loading cloudName: " +
                      os.environ["cloudName"] + " from environment variable.")
            else:
                os.environ["cloudName"] = args["cloud_name"]
        except Exception:
            if not args["cloud_name"]:
                parser.error(
                    "cloud_name parameter is empty. Either Pass the argument -c followed by cloud_name, eg. perfectoai -c demo or add it as a cloudName environment variable"
                )
                exit
            os.environ["cloudName"] = args["cloud_name"]
        try:
            if not args["security_token"]:
                print("Loading securityToken: " +
                      os.environ["securityToken"] + " from environment variable.")
            else:
                os.environ["securityToken"] = args["security_token"]
        except Exception:
            if not args["security_token"]:
                parser.error(
                    "security_token parameter is empty. Pass the argument -c followed by cloud_name, eg. perfectoai -c demo -s <<TOKEN>> || perfectoai -c demo -s <<user>>:<<password>> or add it as a securityToken environment variable"
                )
                exit
            os.environ["securityToken"] = args["security_token"]
        os.environ[
            "perfecto_logo"
        ] = "https://logo.clearbit.com/www.perfecto.io?size=120"
        if args["logo"]:
            if str("www.").lower() not in str(args["logo"]).lower():
                raise Exception(
                    "Kindly provide valid client website url. Sample format: www.perfecto.io"
                )
                sys.exit(-1)
            new_logo = "https://logo.clearbit.com/" + \
                args["logo"] + "?size=120"
            validate_logo(new_logo)
            os.environ["company_logo"] = new_logo
        else:
            os.environ["company_logo"] = os.environ["perfecto_logo"]
        if args["debug"]:
            global debug
            debug = "true"
        if args["email"]:
            os.environ["bgcolor"] = "beige"
            if args["bgcolor"]:
                os.environ["bgcolor"] = args["bgcolor"]
            email_report = args["email"]

            try:
                global criteria
                global jobNumber
                global jobName
                global startDate
                global endDate
                global consolidate
                global trends
                global report
                global addInfo
                global tags
                global reportTag
                global live_report_filename
                global recommend_tag
                global title_heading
                global showIssues
                global ci
                addInfo = ""
                tags = ""
                consolidate = ""
                xlformat = "csv"
                port = ""
                orcaport = "8000"
                temp = ""
                regex = ""
                ci_name = ""
                ci_jenkins_url = ""
                ci_username = ""
                ci_token = ""
                ci = "true"
                recommendations_count = 5
                recommend_tag = ""
                report_array = email_report.split("|")
                title_heading = str(
                    os.environ["cloudName"]).upper() + " Report"
                os.environ["title"] = title_heading
                showIssues = "true"
                os.environ["showIssues"] = showIssues
                os.environ["ci_execution"] = ci
                for item in report_array:
                    if "report" in item:
                        report, criteria = get_report_details(
                            item, temp, "report", criteria
                        )
                    if "jobName" in item:
                        jobName, criteria = get_report_details(
                            item, temp, "jobName", criteria
                        )
                    if "jobNumber" in item:
                        jobNumber, criteria = get_report_details(
                            item, temp, "jobNumber", criteria
                        )
                    if "startDate" in item:
                        startDate, criteria = get_report_details(
                            item, temp, "startDate", criteria
                        )
                    if "endDate" in item:
                        endDate, criteria = get_report_details(
                            item, temp, "endDate", criteria
                        )
                    if "consolidate" in item:
                        consolidate, criteria = get_report_details(
                            item, temp, "consolidate", criteria
                        )
                    if "xlformat" in item:
                        xlformat, criteria = get_report_details(
                            item, temp, "xlformat", criteria
                        )
                    if "port" in item:
                        port, criteria = get_report_details(
                            item, temp, "port", criteria
                        )
                    if "orcaport" in item:
                        orcaport, criteria = get_report_details(
                            item, temp, "orcaport", criteria
                        )
                    if "trends" in item:
                        trends, criteria = get_report_details(
                            item, temp, "trends", criteria
                        )
                    if "addInfo" in item:
                        addInfo, criteria = get_report_details(
                            item, temp, "addInfo", criteria
                        )
                    if "tags" in item:
                        tags, criteria = get_report_details(
                            item, temp, "tags", criteria
                        )
                    if "reportTag" in item:
                        reportTag, criteria = get_report_details(
                            item, temp, "reportTag", criteria
                        )
                    if "attachmentName" in item:
                        live_report_filename, criteria = get_report_details(
                            item, temp, "attachmentName", criteria
                        )
                    if "regex" in item:
                        regex, criteria = get_report_details(
                            item, temp, "regex", criteria
                        )
                    if "ci_name" in item:
                        ci_name, criteria = get_report_details(
                            item, temp, "ci_name", criteria
                        )
                    if "ci_jenkins_url" in item:
                        ci_jenkins_url, criteria = get_report_details(
                            item, temp, "ci_jenkins_url", criteria
                        )
                    if "ci_username" in item:
                        ci_username, criteria = get_report_details(
                            item, temp, "ci_username", criteria
                        )
                    if "ci_token" in item:
                        ci_token, criteria = get_report_details(
                            item, temp, "ci_token", criteria
                        )
                    if "recommendations" in item:
                        recommendations_count, criteria = get_report_details(
                            item, temp, "recommendations", criteria
                        )
                    if "recommend_tag" in item:
                        recommend_tag, criteria = get_report_details(
                            item, temp, "recommend_tag", criteria
                        )
                    if "title" in item:
                        title_heading, criteria = get_report_details(
                            item, temp, "title", criteria
                        )
                    if "showIssues" in item:
                        showIssues, criteria = get_report_details(
                            item, temp, "showIssues", criteria
                        )
                    if "ci" in item:
                        ci, criteria = get_report_details(
                            item, temp, "ci", criteria
                        )

            except Exception as e:
                raise Exception(
                    "Verify parameters of report, split them by | seperator/ " +
                    str(e)
                )
                sys.exit(-1)

            if "attachmentName=" in email_report:
                live_report_filename = live_report_filename + ".html"
            os.environ["title"] = title_heading
            os.environ["xlformat"] = xlformat
            os.environ["regex"] = ""
            os.environ["regex"] = regex
            os.environ["consolidate"] = ""
            os.environ["consolidate"] = consolidate
            os.environ["orcaport"] = orcaport
            os.environ["ci_name"] = ""
            os.environ["ci_name"] = ci_name
            os.environ["ci_jenkins_url"] = ""
            os.environ["ci_jenkins_url"] = ci_jenkins_url
            os.environ["ci_username"] = ""
            os.environ["ci_username"] = ci_username
            os.environ["ci_token"] = ""
            os.environ["ci_token"] = ci_token
            os.environ["recommendations"] = str(recommendations_count)
            os.environ["recommend_tag"] = ""
            os.environ["recommend_tag"] = str(recommend_tag)
            os.environ["showIssues"] = str(showIssues)
            os.environ["ci_execution"] = str(ci)

            if os.name == 'nt':
                DETACHED_PROCESS = 0x00000008
                subprocess.call('taskkill /F /IM Electron.exe',
                                creationflags=DETACHED_PROCESS)
                orcaport = os.environ["orcaport"]
                subprocess.Popen(["orca", "serve", "-p", orcaport],
                                 stdout=subprocess.PIPE, shell=True)
                io.orca.config.server_url = "http://localhost:" + orcaport
                io.orca.status._props["state"] = "validated"

            filelist = glob.glob(os.path.join("*." + xlformat))
            for f in filelist:
                os.remove(f)
            filelist = glob.glob(os.path.join("*.txt"))
            for f in filelist:
                os.remove(f)
            filelist = glob.glob(os.path.join("*.html"))
            for f in filelist:
                os.remove(f)

            graphs, interactive_graphs, df = prepareReport(
                jobName, jobNumber, reportTag)
            if jobName:
                criteria += "Job: " + jobName.replace(";", "; ") + "<br>"
            if jobNumber != "":
                criteria += "Build Number: " + \
                    jobNumber.replace(";", "; ") + "<br>"
            if os.environ["consolidate"] != "":
                criteria += (
                    "start: "
                    + str(df["startTime"].iloc[-1]).split(" ", 1)[0]
                    + ", end: "
                    + str(df["startTime"].iloc[0]).split(" ", 1)[0]
                )
            elif startDate != "":
                if "-" not in startDate:
                    criteria += "Start: " + str(datetime.strptime(str(datetime.fromtimestamp(int(int(startDate) / 1000))), "%Y-%m-%d %H:%M:%S")) + \
                        ", End: " + \
                        str(datetime.strptime(str(datetime.fromtimestamp(
                            int(int(endDate) / 1000))), "%Y-%m-%d %H:%M:%S"))
                else:
                    criteria += "Start: " + startDate + ", End: " + endDate
            global title
            title = ""
            if addInfo != "" or tags != "":
                title = report + criteria + "<br>Info: " + addInfo + "<br>" + tags
            else:
                title = report + criteria
            if reportTag != "":
                title += "<br> tags: " + reportTag
            execution_summary = {}
            execution_summary = create_pie(df, "", "status", "device_summary",)
            failed = df[(df["status"] == "FAILED")]
            passed = df[(df["status"] == "PASSED")]
            blocked = df[(df["status"] == "BLOCKED")]
            failed_blocked = df[
                (df["status"] == "FAILED") | (df["status"] == "BLOCKED")
            ]

            # monthly stats
            df["platforms/0/deviceType"] = df["platforms/0/deviceType"].fillna(
                "Others")
            df["platforms/0/os"] = df["platforms/0/os"].fillna("Others")
            df = df.rename(
                columns={
                    "platforms/0/deviceType": "Platform",
                    "platforms/0/os": "OS",
                    "status": "Test Status",
                    "failureReasonName": "Custom Failure Reason",
                    "platforms/0/osVersion": "Version"
                }
            )
            monthlyStats = {}
            monthlyStats = df.pivot_table(
                index=["month", "week", "Platform", "OS", "Version"],
                columns="Test Status",
                values="name",
                aggfunc="count",
                margins=True,
                fill_value=0,
            ).fillna("")
            for column in monthlyStats.columns:
                monthlyStats[column] = (
                    monthlyStats[column].astype(
                        str).replace("\.0", "", regex=True)
                )
            monthlyStats = monthlyStats.to_html(
                classes="mystyle",
                table_id="report",
                index=True,
                render_links=True,
                escape=False,
            )
            monthlyStats = style_df_email(monthlyStats)
            failurereasons = {}
            if "Custom Failure Reason" not in df.columns:
                df["Custom Failure Reason"] = ""
            df.loc[(df["Custom Failure Reason"].astype(str) == "unclassified error") & (
                df["Test Status"].astype(str) == "PASSED"), "Custom Failure Reason"] = NaN
            failurereasons = crosstab(
                df["Custom Failure Reason"], df["Test Status"]
            )
            failurereasons = failurereasons.to_html(
                classes="mystyle",
                table_id="report",
                index=True,
                render_links=True,
                escape=False,
            )
            failurereasons = style_df_email(failurereasons)
            # top failed TCs
            failed = failed[~failed.name.isin(["Interactive session"])]
            topfailedTCNames = (
                failed.groupby(["name"])
                .size()
                .reset_index(name="#Failed")
                .sort_values("#Failed", ascending=False)
                .head(5)
            )
            reportURLs = []
            for ind in topfailedTCNames.index:
                reportURLs.append(
                    failed.loc[
                        failed["name"] == topfailedTCNames["name"][ind], "reportURL"
                    ].iloc[0]
                )
            topfailedTCNames["Result"] = reportURLs
            topfailedTCNames["Result"] = topfailedTCNames["Result"].apply(
                lambda x: "{0}".format(x)
            )
            for ind in topfailedTCNames.index:
                topfailedTCNames.loc[topfailedTCNames["name"].index == ind, "name"] = (
                    '<a target="_blank" href="'
                    + topfailedTCNames["Result"][ind]
                    + '">'
                    + topfailedTCNames["name"][ind]
                    + "</a>"
                )
            topfailedTCNames = topfailedTCNames.drop("Result", 1)
            topfailedTCNames.columns = ["Top Failed Tests", "#Failed"]
            topfailedtable = {}
            topfailedtable = topfailedTCNames.to_html(
                classes="mystyle",
                table_id="report",
                index=False,
                render_links=True,
                escape=False,
            )
            topfailedtable = style_df_email(topfailedtable)
            process_failures(df, failed_blocked, topfailedTCNames,
                             graphs, interactive_graphs, failed, passed, blocked, failurereasons, monthlyStats, topfailedtable, execution_summary)
            from http.server import SimpleHTTPRequestHandler
            import socket
            from socketserver import TCPServer
            import webbrowser

            if port != "":
                PORT = int(port)
                url = (
                    "http://"
                    + socket.gethostbyname(socket.gethostname())
                    + ":"
                    + str(PORT)
                    + "/"
                    + live_report_filename
                )
                print("Live dashboard url: " + url)
                with TCPServer(("", PORT), SimpleHTTPRequestHandler) as httpd:
                    print("serving at port", PORT)
                    webbrowser.open(url, new=0)
                    httpd.serve_forever()
            else:
                if("false" in str(os.environ["ci_execution"])):
                    webbrowser.open(
                        "file://" + os.path.join(os.getcwd(), live_report_filename), new=0
                    )
                print(
                    "Interactive Report: file://"
                    + os.path.join(os.getcwd(), live_report_filename)
                )
                print(
                    "Emailable Report: file://"
                    + os.path.join(os.getcwd(), email_report_filename)
                )
                end = datetime.now().replace(microsecond=0)
                print("Total Time taken:" + str(end - start_time))
    except Exception as e:
        if os.name == 'nt':
            DETACHED_PROCESS = 0x00000008
            subprocess.call('taskkill /F /IM Electron.exe',
                            creationflags=DETACHED_PROCESS)
        raise Exception("Oops!", e)


if __name__ == "__main__":
    main()
    if os.name == 'nt':
        DETACHED_PROCESS = 0x00000008
        subprocess.call('taskkill /F /IM Electron.exe',
                        creationflags=DETACHED_PROCESS)
    sys.exit()
