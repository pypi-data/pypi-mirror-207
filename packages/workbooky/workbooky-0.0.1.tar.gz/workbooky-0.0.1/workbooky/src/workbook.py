"""Methods to validate and open workbook and worksheets."""
import os
from openpyxl import load_workbook, utils as openpyxl_utils
import zipfile
from termcolor import cprint
import colorama

colorama.init()

ERROR_COLOUR = 'red'

class Workbook():
    def __init__(self, workbook_path: str):
        self.path = str(workbook_path)
        self.workbook = self._get_workbook()

    def _get_workbook(self) -> object:
        # Return the workbook if it exists, otherwise quit.
        if not os.path.isfile(self.path):
            cprint(f"Cannot find workbook in: {self.path}", ERROR_COLOUR)
            cprint(f"Is it named correctly?", ERROR_COLOUR)
            quit()
        try:
            workbook = load_workbook(filename=self.path, data_only=True)
            return workbook
        except openpyxl_utils.exceptions.InvalidFileException:
            cprint(f"{self.path} is not a valid excel file", ERROR_COLOUR)
            quit()
        except zipfile.BadZipFile:
            cprint(f"{self.path} is not a valid excel file", ERROR_COLOUR)
            quit()

    def get_worksheet(self, sheet_name: str) -> object:
        # Return the worksheet if it exists, otherwise quit.
        try:
            worksheet = self.workbook[sheet_name]
            return worksheet
        except KeyError:
            cprint(f"There is no {sheet_name} sheet in the workbook {self.path}", ERROR_COLOUR)
            quit()

    def save(self):
        # Save the workbook
        self.workbook.save()
