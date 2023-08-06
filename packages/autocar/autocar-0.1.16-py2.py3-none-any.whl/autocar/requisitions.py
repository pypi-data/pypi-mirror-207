from pathlib import Path
from datetime import datetime
from shutil import copy, move
import xlwings as xw
import openpyxl as pyxl
from openpyxl.styles import Alignment

xl_app = xw.apps.active

# Set workbook variables
wb: xw.Book = xw.Book.caller()
ws_input: xw.Sheet = wb.sheets['Input']
ws_settings: xw.Sheet = wb.sheets['Settings']
network_save = Path(ws_settings['c_networkPath'].value)


def button_process(str_path: str):
    # Declare variables
    this_path = Path(str_path)
    files_not_opened = []
    files_to_move = []
    files_not_moved = []
    files_old_format = []
    # Get unique identifier
    time_identifier = datetime.now().strftime('%m%d%y-%H%M%S')
    i_identifier = 1

    (this_path / 'TODO').mkdir(exist_ok=True)
    (this_path / 'Complete').mkdir(exist_ok=True)

    for path in this_path.glob('*.xlsx'):
        try:
            po_identifier = 'PR-' + time_identifier + '-' + str(i_identifier)
            change_file(path, po_identifier, this_path)
            files_to_move.append(po_identifier + '.xlsx')
            i_identifier += 1
        except Exception:
            files_not_opened.append(path.stem)

    for path in files_to_move:
        try:
            copy((this_path / 'TODO') / path, Path(network_save) / path)
        except Exception:
            files_not_moved.append(path)

    for path in this_path.glob('*.xlsx'):
        try:
            move(path, (this_path / 'Complete') / path.name)
        except Exception:
            files_not_moved.append(path)

    for path in this_path.glob('*.xls'):
        files_old_format.append(path)

    if files_not_opened:
        xl_app.alert(' '.join(map(str, files_not_opened)), title='Files not opened', mode='critical')

    if files_not_moved:
        xl_app.alert(' '.join(map(str, files_not_moved)), title='Files not moved', mode='critical')

    if files_old_format:
        xl_app.alert('Change the format of these files:\n' + ' '.join(map(str, files_old_format)), title='Old format', mode='critical')

    xl_app.alert('Complete.', title='Complete', mode='info')


def change_file(path, po_identifier, this_path):
    pr_wb = pyxl.load_workbook(filename=path, data_only=True)
    ws = pr_wb.active

    ws['C1'].value = po_identifier
    ws['C1'].alignment = Alignment(wrap_text=False)
    pr_wb.save((this_path / 'TODO') / (po_identifier + '.xlsx'))


if __name__ == "__main__":
    pass
