
import os, copy, pickle
import tkinter as tk
from openpyxl import (load_workbook, Workbook)
import tksheet

colors=['#FFBBBB', '#BBFFBB', '#BBBBFF', '#CCCCFF']
path4xlsx=''


class sheet:

    def saveCleanedData(self, evnt=None):
        wb=Workbook()
        ws = wb.active
        headers = self.shtL.headers()
        data = self.shtL.get_sheet_data()

        if headers:
            for c, v in enumerate(headers):
                ws[f'{chr(65+c)}1']=v
            for r, row in enumerate(data):
                for c, v in enumerate(row):
                    ws[f'{chr(65+c)}{r+2}']=v
        else:
            for r, row in enumerate(data):
                for c, v in enumerate(row):
                    ws[f'{chr(65+c)}{r+1}']=v

        fn = self.title.split('.xlsx')[0]+'_cleaned.xlsx'
        try:
            wb.save(fn)
            self.hL.config(text=f'{fn}: saving is completed.', fg='green')
        except PermissionError:
            self.hL.config(text=f'{fn} is open. Try again.', fg='red')


    def combineRowsColumns(self, evnt=None):
        if not self.shtL.anything_selected(): return

        rowFrom, colFrom, rowTo, colTo = self.shtL.get_all_selection_boxes()[0]
        data = self.shtL.get_sheet_data()
        self.dataBakup.append(copy.deepcopy(data))
        
        if rowTo-rowFrom == len(data) and colTo-colFrom < len(data[0]):
            self.actionDone=['combineColumns', [colFrom, colTo]]

            for row, dat in enumerate(data):
                tmp=[s for s in [s.strip() if isinstance(s, str) else str(s) for s in dat[colFrom:colTo]] if s != 'None']
                data[row]=dat[:colFrom]+[' '.join(tmp)]+dat[colTo:]

        elif colTo-colFrom == len(data[0]) and rowTo-rowFrom <len(data):
            self.actionDone=['combineRows', [rowFrom, rowTo]]

            row=[]
            for j in range(len(data[rowFrom])):
                column=[d[j].strip() if isinstance(d[j], str) else str(d[j]) for d in data[rowFrom:rowTo]]
                row.append(' '.join([s for s in column if s != 'None']))                
                        
            data=data[:rowFrom]+[row]+data[rowTo:]
            
        print(self.actionDone)
        if self.option2SaveActions: self.saveAction2File()

        self.shtL.set_sheet_data(data)
        self.highlightRows()


    def addCells(self, evnt=None):
        if not self.shtL.anything_selected(): return

        rowFrom, col, _, _ = self.shtL.get_all_selection_boxes()[0]
        data = self.shtL.get_sheet_data()
        self.dataBakup.append(copy.deepcopy(data))
        self.actionDone=['add', [col, rowFrom]]

        for row in data:
            row.append(None)
        data[rowFrom].pop()
        data[rowFrom].insert(col, None)

        data[0][-1]='Xtra'

        print(self.actionDone)
        if self.option2SaveActions: self.saveAction2File()
        
        self.shtL.set_sheet_data(data)
        if self.option4Repeat: self.checkSimilars()
        self.highlightRows()



    def saveAction2File(self):

        txt, lst = self.actionDone
        if txt == 'delRows':
            with open('actions.txt', 'a', encoding='utf-8') as f:
                f.write(f'#{self.actionDone}\n')
                f.write(f'self.shtL.create_selection_box({lst[0]},0,{lst[1]},len(self.shtL.get_row_data(0)))\n')
                f.write('self.shtL.recreate_all_selection_boxes()\n')
                f.write('self.delRowsColumns()\n\n')

        elif txt == 'delColumns':
            with open('actions.txt', 'a', encoding='utf-8') as f:
                f.write(f'#{self.actionDone}\n')
                f.write('data = self.shtL.get_sheet_data()\n')
                f.write(f'data=[row[:{lst[0]}]+row[{lst[1]}:] for row in data]\n')
                f.write('self.shtL.set_sheet_data(data)\n')
                f.write('self.shtL.set_all_cell_sizes_to_text(redraw = True)\n')
                f.write('self.highlightRows()\n\n')

        elif txt == 'combineRows':
            with open('actions.txt', 'a', encoding='utf-8') as f:
                f.write(f'#{self.actionDone}\n')
                f.write('data = self.shtL.get_sheet_data()\n')
                f.write('row=[]\n')
                f.write(f"for j in range(len(data[{lst[0]}])):\n")
                f.write(f"    column=[]\n")
                f.write(f"    for d in data[{lst[0]}:{lst[1]}]:\n")
                f.write(f"        if isinstance(d[j], str):\n")
                f.write(f"            column.append(d[j].strip())\n")
                f.write(f"        else:\n")
                f.write(f"            column.append(str(d[j]))\n")
                f.write("    row.append(' '.join([s for s in column if s != 'None']))\n")
                f.write(f"data=data[:{lst[0]}]+[row]+data[{lst[1]}:]\n")
                f.write('self.shtL.set_sheet_data(data)\n')
                f.write('self.shtL.set_all_cell_sizes_to_text(redraw = True)\n')
                f.write('self.highlightRows()\n\n')


        elif txt == 'combineColumns':
            with open('actions.txt', 'a', encoding='utf-8') as f:
                f.write(f'#{self.actionDone}\n')
                f.write('data = self.shtL.get_sheet_data()\n')
                f.write('for row, dat in enumerate(data):\n')
                f.write(f"    tmp=[s for s in [s.strip() if isinstance(s, str) else str(s) for s in dat[{lst[0]}:{lst[1]}]] if s != 'None']\n")
                f.write(f"    data[row]=dat[:{lst[0]}]+[' '.join(tmp)]+dat[{lst[1]}:]\n")
                f.write('self.shtL.set_sheet_data(data)\n')
                f.write('self.shtL.set_all_cell_sizes_to_text(redraw = True)\n')
                f.write('self.highlightRows()\n\n')

        elif txt == 'add':
            with open('actions.txt', 'a', encoding='utf-8') as f:
                f.write(f'#{self.actionDone}\n')
                f.write(f'self.shtL.select_cell({lst[1]}, {lst[0]})\n')
                f.write('self.addCells()\n\n')

        elif txt == 'sub':
            with open('actions.txt', 'a', encoding='utf-8') as f:
                f.write(f'#{self.actionDone}\n')
                f.write(f'self.shtL.select_cell({lst[1]}, {lst[0]})\n')
                f.write('self.subCells()\n\n')




    def subCells(self, evnt=None):
        if not self.shtL.anything_selected(): return

        rowFrom, col, _, _ = self.shtL.get_all_selection_boxes()[0]

        data = self.shtL.get_sheet_data()
        self.dataBakup.append(copy.deepcopy(data))
        self.actionDone=['sub', [col, rowFrom]]
        data[rowFrom].append(data[rowFrom].pop(col))
        
        print(self.actionDone)
        if self.option2SaveActions: self.saveAction2File()

        self.shtL.set_sheet_data(data)
        if self.option4Repeat: self.checkSimilars()
        self.highlightRows()
        


    def delRowsColumns(self, evnt=None):
        if not self.shtL.anything_selected(): return

        rowFrom, colFrom, rowTo, colTo = self.shtL.get_all_selection_boxes()[0]

        data = self.shtL.get_sheet_data()
        self.dataBakup.append(copy.deepcopy(data))

        if rowTo-rowFrom < len(data) and colTo-colFrom == len(data[0]):

            if rowTo-rowFrom == 1 and len(self.highlighted_rows)>=1:
                for j in self.highlighted_rows[::-1]:
                    data.pop(j)

                self.actionDone=['delSelectedRows', [None, None]]
                print(self.actionDone)
                if self.option2SaveActions:
                    with open('actions.txt', 'a', encoding='utf-8') as f:
                        f.write(f'#{self.actionDone}\n')
                        f.write('data = self.shtL.get_sheet_data()\n')
                        f.write(f'self.highlighted_rows = {self.highlighted_rows}\n')
                        f.write('for j in self.highlighted_rows[::-1]:\n')
                        f.write("    data.pop(j)\n\n")
                        f.write('self.shtL.set_sheet_data(data)\n')
                        f.write('self.shtL.set_all_cell_sizes_to_text(redraw = True)\n')
                        f.write('self.highlightRows()\n\n')

                self.shtL.set_sheet_data(data)
                self.highlightRows()
            else:
                self.actionDone=['delRows', [rowFrom, rowTo]]
                data=data[:rowFrom]+data[rowTo:]

                print(self.actionDone)
                if self.option2SaveActions: self.saveAction2File()

                self.shtL.set_sheet_data(data)
                if self.option4Repeat: self.checkSimilars()
                self.highlightRows()

        elif rowTo-rowFrom == len(data) and colTo-colFrom < len(data[0]):
            self.actionDone=['delColumns', [colFrom, colTo]]
            data=[row[:colFrom]+row[colTo:] for row in data]

            print(self.actionDone)
            if self.option2SaveActions: self.saveAction2File()

            self.shtL.set_sheet_data(data)
            self.highlightRows()

        
        
    def repeat(self, evnt=None):
        data = self.shtL.get_sheet_data()
        if self.actionDone[0] == 'delRows':
            rowsN, rows2delete=self.actionDone[2]
            if rowsN > 1:
                for j in rows2delete[::-1]:
                    data=data[:j]+data[j+rowsN:]
            else:
                print(rows2delete, len(self.kinds), len(data))
                for j in rows2delete[::-1]:
                    data.pop(j)
            
        elif self.actionDone[0] == 'sub':
            col, rows2delete=self.actionDone[2]
            for j in rows2delete:
                data[j].append(data[j].pop(col))

        elif self.actionDone[0] == 'add':
            col, rows2add=self.actionDone[2]
            for j in rows2add:
                data[j].insert(col, None)
                data[j].pop()

        self.shtL.set_sheet_data(data)
        self.highlightRows()
            

    
    def checkSimilars(self):
        data = self.shtL.get_sheet_data()
        flag2Repeat=False
        if self.actionDone[0] == 'delRows':
            rowFrom, rowTo = self.actionDone[1]
            if rowTo-rowFrom > 1:
                deletedRows=[row[0] for row in self.dataBakup[-1][rowFrom:rowTo]]
                rowsN=len(deletedRows)
                rows2delete=[]
                for j in range(len(data)):
                    if  [row[0] for row in data[j:j+rowsN]] == deletedRows:
                        rows2delete.append(j)
            else:
                deletedRows = self.dataBakup[-1][rowFrom]

                tmpX=[]
                for itemX in deletedRows:
                    if itemX is None:
                        kindX = 'none'
                    elif isinstance(itemX, (int, float)):
                        kindX = 'num'
                    elif isinstance(itemX, str):
                        if itemX.isalpha():
                            kindX = 'alpha'
                        elif itemX.replace('.','').isnumeric():
                            kindX = 'num'
                        else:
                            kindX = 'alnum'
                    else:
                        kindX='???'
                    tmpX.append(kindX)
                rowKind=tuple(tmpX)
                
                rowsN=1
                rows2delete=[]
                for j, kind in enumerate(self.kinds):
                    if j != rowFrom and kind == rowKind:
                        rows2delete.append(j)
                
            if rows2delete:
                self.actionDone.append((rowsN, rows2delete))
                
                flag2Repeat=True

        elif self.actionDone[0] == 'sub':
            col, rowFrom = self.actionDone[1]

            deletedRowTypes=[self.getType(item) for item in self.dataBakup[-1][rowFrom]]
            rows2delete=[]
            for j, row in enumerate(data[1:]):
                if [self.getType(item) for item in row] == deletedRowTypes:
                    rows2delete.append(j+1)
                    
            if rows2delete:
                self.actionDone.append((col, rows2delete))
                flag2Repeat=True
            
        elif self.actionDone[0] == 'add':
            col, rowFrom = self.actionDone[1]

            addedRowTypes=[self.getType(item) for item in self.dataBakup[-1][rowFrom]]+[None]
            rows2add=[]
            for j, row in enumerate(data[1:]):
                if [self.getType(item) for item in row] == addedRowTypes:
                    rows2add.append(j+1)

            if rows2add:
                self.actionDone.append((col, rows2add))
                flag2Repeat=True

        if flag2Repeat: self.repeat()


    def highlightRows(self):
        self.shtL.set_all_cell_sizes_to_text(redraw = True)
        rows = self.shtL.get_sheet_data()

        self.kinds=[]
        for row in rows:
            tmp=[]
            for item in row:
                if item is None:
                    kind='none'
                elif isinstance(item, (int, float)):
                    kind='num'
                elif isinstance(item, str):
                    if item.isalpha():
                        kind='alpha'
                    elif item.replace('.','').isnumeric():
                        kind='num'
                    else:
                        kind = 'alnum'
                else:
                    kind='???'
                tmp.append(kind)
            self.kinds.append(tuple(tmp))

        kindsX=self.kinds[1:]
        tmp={kind: kindsX.count(kind) for kind in set(kindsX)}
        patterns={kind: count for kind, count in sorted(tmp.items(), key=lambda t: t[1])}

        patternlist=iter(patterns.keys())
        indices0=[j for j, p in enumerate(next(patternlist)) if p == 'none']

        while indices0 and patternlist:
            try:
                patt = next(patternlist)
            except:
                break

            indicesJ =[j for j, p in enumerate(patt) if p == 'none']
            for v in copy.deepcopy(indices0):
                if v not in indicesJ: indices0.remove(v)

        if self.option4DeleteColsSameValue:
            content1=list(enumerate(rows[1]))
            for row in rows[2:]:
                contentRow = list(enumerate(row))
                for c1 in content1:
                    if contentRow[c1[0]][1] != c1[1]:
                        content1.remove(c1)
                if not content1: break
            
        self.shtL.dehighlight_all()
        if indices0:
            print('deleting column of ', indices0[-1], ' for all empty')
            if self.option2SaveActions:
                with open('actions.txt', 'a', encoding='utf-8') as f:
                    f.write(f'#deleting column of {indices0[-1]} for all empty\n')

            colFrom=indices0[-1]
            rows=[row[:colFrom]+row[colFrom+1:] for row in rows]
            self.shtL.set_sheet_data(rows)

        elif self.option4DeleteColsSameValue and content1:
            print('deleting column of ', content1[-1][0], ' for all same')
            if self.option2SaveActions:
                with open('actions.txt', 'a', encoding='utf-8') as f:
                    f.write(f'#deleting column of {indices0[-1]} for all same text\n')

            colFrom=content1[-1][0]
            rows=[row[:colFrom]+row[colFrom+1:] for row in rows]
            self.shtL.set_sheet_data(rows)

        else:
            indices = [j for j, kind in enumerate(kindsX) if kind == list(patterns.keys())[0]]
            self.highlighted_rows=indices
            self.shtL.highlight_rows(rows = indices, bg = colors[2], redraw = True)
            self.shtL.select_row(row=indices[0], redraw=True)
            self.shtL.see(row=indices[0])
                
        self.saveCleanedData()
        if self.hL['fg'] == 'green':
            self.hL.config(text=f'cells = {len(rows[0])}×{len(rows)}   row patterns = {len(patterns)}')



    @staticmethod
    def getType(a):
        if isinstance(a, str):
            if a.isnumeric() or ('.' in a and a.replace('.','').isnumeric()): return 'num'
            else: return 'str'
        elif isinstance(a, int):
            return 'int'
        elif isinstance(a, float):
            return 'float'
        

    def restore(self, event=None):
        if not self.dataBakup:
            data = self.shtL.get_sheet_data()
            self.hL.config(text=f'No more data to restore. cells = {len(data[0])}×{len(data)}', fg='red')
            return
            
        data=self.dataBakup.pop()
        self.shtL.set_sheet_data(data)
        self.shtL.set_all_cell_sizes_to_text(redraw = True)
        self.shtL.dehighlight_all()
        self.actionDone=None
        self.kinds=[]
        self.hL.config(text=f'cells = {len(data[0])}×{len(data)}', fg='black')


    def options(self, event=None):

        def update(e=None):
            self.option2SaveActions=iAction.get()
            if self.option2SaveActions:
                if not os.path.exists('actions.txt'):
                    with open('actions.txt', 'w', encoding='utf-8') as f:
                        f.write(f'#{self.title}\n')

            self.option4Repeat = iRepeat2.get()
            self.option4DeleteColsSameValue = iRepeat.get()
            pop.destroy()
            pop.grab_release()

            
        pop=tk.Toplevel(self.hw)
        pop.title('Options')

        Fr0=tk.Frame(pop, highlightbackground="blue", highlightthickness=1)
        Fr0.pack(pady=5, padx=10, ipadx=10)

        tk.Label(Fr0, text='Repeat delete/add process automatically:').pack(side='left', padx=10)

        iRepeat2=tk.IntVar(Fr0)
        tk.Radiobutton(Fr0, text="No", variable=iRepeat2, value=0).pack(side='left', padx=10)
        tk.Radiobutton(Fr0, text="Yes", variable=iRepeat2, value=1).pack(side='left')
        iRepeat2.set(self.option4Repeat)

        Fr1=tk.Frame(pop, highlightbackground="blue", highlightthickness=1)
        Fr1.pack(pady=5, padx=10, ipadx=10)

        tk.Label(Fr1, text='Delete columns with the same value:').pack(side='left', padx=10)

        iRepeat=tk.IntVar(Fr1)
        tk.Radiobutton(Fr1, text="No", variable=iRepeat, value=0).pack(side='left', padx=10)
        tk.Radiobutton(Fr1, text="Yes", variable=iRepeat, value=1).pack(side='left')
        iRepeat.set(self.option4DeleteColsSameValue)
        
        Fr4=tk.Frame(pop, highlightbackground="blue", highlightthickness=1)
        Fr4.pack(pady=5, padx=10, ipadx=10)

        tk.Label(Fr4, text='Save actions to text file:').pack(side='left', padx=10)

        iAction=tk.IntVar(Fr4)
        tk.Radiobutton(Fr4, text="No", variable=iAction, value=0).pack(side='left', padx=10)
        tk.Radiobutton(Fr4, text="Yes", variable=iAction, value=1).pack(side='left')
        iAction.set(self.option2SaveActions)

        tk.Button(pop, text='Update and Exit', command=update).pack(pady=10)
        
        pop.wait_visibility()
        pop.grab_set()
        


    def combineRowsSpecial(self, evnt=None):

        def update(event=None):
            rows = self.shtL.get_sheet_data()
            self.dataBakup.append(copy.deepcopy(rows))
            
            selectedRowTo=self.highlighted_rows[1:]
            selectedRowTo.append(len(rows))

            colsN=len(rows[0])
            combinedData=rows[:self.highlighted_rows[0]]
            for from_, to in zip(self.highlighted_rows, selectedRowTo):
                row=[]
                for jj in range(colsN):
                    tmp=' '.join([f'{row[jj]}' for row in rows[from_:to] if row[jj]])
                    if tmp.replace('.','').isnumeric():
                        row.append(eval(tmp))
                    else:
                        row.append(tmp)

                combinedData.append(row)

            self.actionDone=['combineRowsSpecial', [None, None]]

            if self.option2SaveActions:
                with open('actions.txt', 'a', encoding='utf-8') as f:
                    f.write(f'#{self.actionDone}\n')
                    f.write('rows = self.shtL.get_sheet_data()\n')
                    f.write(f'self.highlighted_rows = {self.highlighted_rows}\n')
                    f.write('selectedRowTo = self.highlighted_rows[1:]\n')
                    f.write('selectedRowTo.append(len(rows))\n')
                    f.write('colsN=len(rows[0])\n')
                    f.write('combinedData=rows[:self.highlighted_rows[0]]\n')
                    f.write('for from_, to in zip(self.highlighted_rows, selectedRowTo):\n')
                    f.write("    row=[]\n")
                    f.write("    for jj in range(colsN):\n")
                    f.write("        tmp=''\n")
                    f.write("        for rowrow in rows[from_:to]:\n")
                    f.write("            if rowrow[jj]: tmp += ' '+f'{rowrow[jj]}'\n\n")
                    f.write("        if tmp.replace('.','').isnumeric():\n")
                    f.write("            row.append(eval(tmp))\n")
                    f.write("        else:\n")
                    f.write("            row.append(tmp)\n\n")
                    f.write("    combinedData.append(row)\n\n")
                    f.write('self.shtL.set_sheet_data(combinedData)\n')
                    f.write('self.shtL.set_all_cell_sizes_to_text(redraw = True)\n')
                    f.write('self.highlightRows()\n\n')

            self.shtL.set_sheet_data(combinedData)
            self.shtL.set_all_cell_sizes_to_text(redraw = True)
            self.highlightRows()
            pop.destroy()
            pop.grab_release()


        pop=tk.Toplevel(self.hw)
        pop.title('Combining rows in special mode')

        Fr0=tk.Frame(pop, highlightbackground="blue", highlightthickness=1)
        Fr0.pack(pady=5, padx=10, ipadx=10)

        msg='The selected rows are:\n'
        msg +=f'{[j+1 for j in self.highlighted_rows]}\n'
        msg +="\nWe're about to combine each row listed above with the rows that follow it."
        msg +="\n\nClick button 'Update', if you agree."
        tk.Label(Fr0, text=msg).pack(side='left', padx=10)

        tk.Button(pop, text='Update and Exit', command=update).pack(pady=10)
        
        pop.wait_visibility()
        pop.grab_set()




    def delRowsSpecial(self, evnt=None):

        rowFrom, _, _, _ = self.shtL.get_all_selection_boxes()[0]
        rows = self.shtL.get_sheet_data()

        def updateIt(event=None):

            j2check=iColumn.get()
            startingString=columnValue.get()
            if not startingString:
                self.hL.config(text='No string is provided yet.', fg='red')
                pop.destroy()
                pop.grab_release()
                return
                
            self.dataBakup.append(copy.deepcopy(rows))
            self.actionDone=['delRowSpecial', [rowFrom, None]]

            rows2delete=[]
            for j, row in enumerate(rows):
                if not row[j2check]: continue
                if isinstance(row[j2check], str):
                    if row[j2check].startswith(startingString):
                        rows2delete.append(j)

            if rows2delete:
                for j in rows2delete[::-1]:
                    rows.pop(j)

                if self.option2SaveActions:
                    with open('actions.txt', 'a', encoding='utf-8') as f:
                        f.write('#Special row deletion\n')
                        f.write('rows = self.shtL.get_sheet_data()\n')
                        f.write(f'rows2delete = {rows2delete}\n')
                        f.write('for j in rows2delete[::-1]: rows.pop(j)\n\n')
                        f.write('self.shtL.set_sheet_data(rows)\n')
                        f.write('self.shtL.set_all_cell_sizes_to_text(redraw = True)\n')
                        f.write('self.highlightRows()\n\n')


            self.shtL.set_sheet_data(rows)
            self.shtL.set_all_cell_sizes_to_text(redraw = True)
            self.highlightRows()
            pop.destroy()
            pop.grab_release()


        def showContent(evnt=None):
            col=iColumn.get()
            wordN=iWordN.get()
            columnValue.delete(0,'end')

            if rows[rowFrom][col] and wordN:
                columnValue.insert(0, rows[rowFrom][col].split(' ')[:wordN])

        
        pop=tk.Toplevel(self.hw)
        pop.option_add('*Font', ('Palatino Linotype',12))
        pop.bind('<Control-s>', updateIt)

        fr0=tk.Frame(pop, highlightbackground="blue", highlightthickness=1)
        fr0.pack(padx=5, pady=5)

        tk.Label(fr0, text='No column to check is provided').pack()
        fr0R=tk.Frame(fr0)
        fr0R.pack(padx=5)
        
        iColumn=tk.IntVar(fr0)
        for j, v in enumerate(rows[rowFrom]):
            tk.Radiobutton(fr0R, text=f'{chr(j+65)}: {v}', variable=iColumn,
                           command=showContent, value=j).pack(side='left', pady=2)

                
        fr1=tk.Frame(pop, highlightbackground="blue", highlightthickness=1)
        fr1.pack(padx=5, pady=5)

        tk.Label(fr1, text='Select the number of words to check:').pack()
        fr1R=tk.Frame(fr1)
        fr1R.pack(padx=5)

        iWordN=tk.IntVar(fr0)
        for j in range(5):
            tk.Radiobutton(fr1R, text=j, variable=iWordN,
                           command=showContent, value=j).pack(side='left', pady=2)
        iWordN.set(2)

        columnValue=tk.Entry(fr1)
        columnValue.pack(pady=(0,10), expand=1, fill='x')

        tk.Button(pop, text='update (Ctrl-s) it and exit', command=updateIt).pack(pady=5)
        
        for j, v in enumerate(rows[rowFrom]):
            if v:
                iColumn.set(j)
                columnValue.insert(0, v.split(' ')[:2])
                columnValue.focus_set()
                break

        pop.wait_visibility()
        pop.grab_set()
        


    def splitCellsSpecial(self, evnt=None):
        rowFrom, _, _, _ = self.shtL.get_all_selection_boxes()[0]
        #print(self.shtL.get_all_selection_boxes())
        data = self.shtL.get_sheet_data()

        def updateIt(event=None):
            j2split = iColumn.get()
            separator = columnValue.get()
            columnHdr = columnHeader.get()

            self.dataBakup.append(copy.deepcopy(data))
            self.actionDone=['splitCellSpecial', [j2split, separator, columnHdr]]

            for row in data[1:]:
                front, rear = row[j2split].split(separator)
                row[j2split] = front.strip()
                row.insert(j2split+1, rear.strip())
                
            data[0].insert(j2split+1, columnHdr)


            if self.option2SaveActions:
                with open('actions.txt', 'a', encoding='utf-8') as f:
                    f.write(f'#{self.actionDone}\n')
                    f.write('data = self.shtL.get_sheet_data()\n')
                    f.write(f'j2split = {j2split}\n')
                    f.write(f"separator = '{separator}'\n")
                    f.write(f"columnHdr = '{columnHdr}'\n")
                    f.write('for row in data[1:]:\n')
                    f.write("    front, rear = row[j2split].split(separator)\n")
                    f.write("    row[j2split] = front.strip()\n")
                    f.write("    row.insert(j2split+1, rear.strip())\n\n")
                    f.write("data[0].insert(j2split+1, columnHdr)\n")
                    f.write('self.shtL.set_sheet_data(data)\n')
                    f.write('self.shtL.set_all_cell_sizes_to_text(redraw = True)\n')
                    f.write('self.highlightRows()\n\n')


            self.shtL.set_sheet_data(data)
            self.shtL.set_all_cell_sizes_to_text(redraw = True)
            self.highlightRows()
            pop.destroy()
            pop.grab_release()


        def showContent(event=None):
            col=iColumn.get()
            columnValue.delete(0,'end')

            word=data[1][col].split(' ')
            for row in data[2:]:
                for w in word[::-1]:
                    if w not in row[col]:
                        word.remove(w)

            columnValue.insert(0, ' '.join(word))

        
        pop=tk.Toplevel(self.hw)
        pop.option_add('*Font', ('Palatino Linotype',12))
        pop.bind('<Control-s>', updateIt)

        fr0=tk.Frame(pop, highlightbackground="blue", highlightthickness=1)
        fr0.pack(padx=5, pady=5)

        tk.Label(fr0, text='Select the column to split.').pack()
        fr0R=tk.Frame(fr0)
        fr0R.pack(padx=5)
        
        iColumn=tk.IntVar(fr0)
        for j, v in enumerate(data[1]):
            tk.Radiobutton(fr0R, text=f'{chr(j+65)}: {v}', variable=iColumn,
                           command=showContent, value=j).pack(side='left', pady=2)

                
        fr1=tk.Frame(pop, highlightbackground="blue", highlightthickness=1)
        fr1.pack(padx=5, pady=5)

        tk.Label(fr1, text='Phrase to split text:').pack()

        columnValue=tk.Entry(fr1)
        columnValue.pack(pady=(0,10), expand=1, fill='x')

        fr2=tk.Frame(pop, highlightbackground="blue", highlightthickness=1)
        fr2.pack(padx=5, pady=5)
        
        tk.Label(fr2, text='Column name for the new column: ').pack(side='left')
        columnHeader=tk.Entry(fr2)
        columnHeader.pack(side='left')
        columnHeader.insert(0, 'UPC')

        tk.Button(pop, text='update it (Ctrl-s) and exit', command=updateIt).pack(pady=5)
        
        pop.wait_visibility()
        pop.grab_set()


    def row_select(self, event=None):
        self.shtL.dehighlight_all()

        if event[0] == 'select_row':
            if not self.kinds: return
            row=event.row
            rowKind=self.kinds[row]
            rowKindCount=self.kinds.count(rowKind)
            if rowKindCount:
                self.highlighted_rows=[j for j, kind in enumerate(self.kinds) if kind==rowKind]
                self.shtL.highlight_rows(rows = self.highlighted_rows, bg = colors[2], redraw = True)

                txt=self.hL['text'].split(' --')[0]
                self.hL.config(text=f'{txt} --  number of rows in same pattern = {len(self.highlighted_rows)}')


    def __init__(self, xlpath: str):

        actions=None
        if os.path.exists('actions.txt'):
            with open('actions.txt', 'r', encoding='utf-8') as f:
                lines=f.readlines()
                actions=''.join(lines)

        self.dataBakup=[]
        self.actionDone=None
        self.option2SaveActions=0
        self.option4Repeat=1
        self.option4DeleteColsSameValue=0
        self.option4SaveInvoice=0
        self.iQtyOld=0
        self.title=xlpath
        self.newItems=[]
        self.updateItems=[]
        self.manualChangeItems=[]
        self.highlighted_rows=[]
        self.kinds = []

        itemHeaders=None
        if os.path.exists(os.path.join(path4xlsx,xlpath)):

            wb=load_workbook(os.path.join(path4xlsx,xlpath), data_only=True)
            data = [[f.value for f in row] for row in wb.worksheets[0].rows]
            data=[row for row in data if row != [None for _ in range(len(data[0]))]]
        else:
            print('Excel file cannot be located')
            return

        
        self.hw=tk.Tk()
        self.hw.option_add('*Font', ('Georgia', 12))
        self.hw.geometry('1000x850')
        self.hw.title(self.title+'     F1: delete cols/rows, F2: special del rows, F3: combine cols/rows, F4: combine special rows, F5: insert cell, F7: delete cell')
        self.hw.bind('<F1>', self.delRowsColumns)
        self.hw.bind('<F2>', self.delRowsSpecial)
        self.hw.bind('<F3>', self.combineRowsColumns)
        self.hw.bind('<F4>', self.combineRowsSpecial)
        self.hw.bind('<F5>', self.addCells)
        self.hw.bind('<F7>', self.subCells)
        self.hw.bind('<F10>', self.splitCellsSpecial)
        
        self.hw.bind('<Control-z>', self.restore)
        self.hw.bind('<Control-o>', self.options)

        self.hL=tk.Label(self.hw, text='...')
        self.hL.pack(pady=2)

        self.shtL = tksheet.Sheet(self.hw)
        self.shtL.pack(expand=1, fill='both')
        self.shtL.enable_bindings()
        self.shtL.disable_bindings(['rc_delete_row', 'rc_delete_column', 'cut', 'paste', 'copy', 'delete',
                                    'rc_insert_row', 'rc_insert_column', 'edit_cell'])
        self.shtL.extra_bindings([('all_select_events', self.row_select)])

        self.shtL.popup_menu_add_command("delete rows/cols (F1)", self.delRowsColumns)
        self.shtL.popup_menu_add_command("combine rows/cols (F3)", self.combineRowsColumns)
        self.shtL.popup_menu_add_command("insert cell (F5)", self.addCells)
        self.shtL.popup_menu_add_command("delete cell (F7)", self.subCells)
        self.shtL.popup_menu_add_command("restore (Ctrl-z)", self.restore)

        self.shtL.set_sheet_data(data)
        self.hL.config(text=f'cells = {len(data[0])}×{len(data)}')

        self.shtL.set_all_cell_sizes_to_text(redraw = True)

        if actions: exec(actions)

        self.hw.mainloop()



if __name__ == "__main__":
    #sheet('AR-93474.xlsx')
    sheet('MG-158427.xlsx')
